import re
import json
import csv
import os
from io import StringIO
from dateutil import parser
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from copy import deepcopy
from datetime import datetime, time
from dtable_events.app.config import EXPORT2EXCEL_DEFAULT_STRING, TIME_ZONE, INNER_DTABLE_DB_URL
from dtable_events.utils import utc_to_tz, get_inner_dtable_server_url, gen_random_option
from dtable_events.utils.constants import ColumnTypes
from dtable_events.utils.geo_location_parser import parse_geolocation_from_tree
from dtable_events.utils.dtable_db_api import DTableDBAPI
from dtable_events.utils.dtable_server_api import DTableServerAPI
from dtable_events.dtable_io.utils import clear_tmp_file, save_file_by_path, get_csv_file, \
    upload_excel_json_add_table_to_dtable_server, append_rows_by_dtable_server, get_related_nicknames_from_dtable, \
    extract_select_options, upload_excel_json_to_dtable_server, get_rows_from_dtable_db, update_rows_by_dtable_db, \
    get_nicknames_from_dtable, get_table_names_by_dtable_server, get_non_duplicated_name
from dtable_events.utils.exception import ExcelFormatError

timezone = TIME_ZONE
VIRTUAL_ID_EMAIL_DOMAIN = '@auth.local'

first_grouped_row_fill = PatternFill(fill_type='solid', fgColor='ffa18b')
second_grouped_row_fill = PatternFill(fill_type='solid', fgColor='ffff4d')
third_grouped_row_fill = PatternFill(fill_type='solid', fgColor='a5f89b')
grouped_row_fills = [first_grouped_row_fill, second_grouped_row_fill, third_grouped_row_fill]


CHECKBOX_TUPLE = (
    ('√', 'x'),
    ('checked', 'unchecked'),
    ('y', 'n'),
    ('yes', 'no'),
    ('enabled', 'disabled'),
    ('on', 'off'),
    ('是', '否'),
    ('完成', '未完成'),
    ('True', 'False'),
    ('true', 'false'),
)
CHECKBOX_STRING_LIST = [string for item in CHECKBOX_TUPLE for string in item]
CHECKBOX_TRUE_LIST = [item[0] for item in CHECKBOX_TUPLE]

# copy from dtable-web/frontend/src/components-form/utils/markdown-utils.js
HREF_REG = r'\[.+\]\(\S+\)|<img src=\S+.+\/>|!\[\]\(\S+\)|<\S+>'
LINK_REG_1 = r'^\[.+\]\((\S+)\)'
LINK_REG_2 = r'^<(\S+)>$'
IMAGE_REG_1 = r'^<img src="(\S+)" .+\/>'
IMAGE_REG_2 = r'^!\[\]\((\S+)\)'

UPDATE_TYPE_LIST = ['number', 'single-select', 'url', 'email', 'text', 'date', 'duration', 'rate', 'checkbox',
                    'multiple-select', 'collaborator']

TEMP_EXPORT_VIEW_DIR = '/tmp/dtable-io/export-view-to-excel/'

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

# image offset in excel cell
FROM_COL_START_OFFSET = 20000
FROM_ROW_START_OFFSET = 20000
TO_COL_START_OFFSET = -80000
TO_ROW_START_OFFSET = -80000
IMAGE_CELL_ROW_HEIGHT = 50
IMAGE_CELL_COLUMN_WIDTH = 30

IMAGE_TMP_DIR = '/tmp/dtable-io/export-excel/images/'

EXPORT_IMAGE_LIMIT = 1000


EXCEL_IMPORT_DIR = '/tmp/dtable-io/'


class EmptyCell(object):
    value = None


def parse_checkbox(cell_value):
    cell_value = str(cell_value)
    return True if cell_value in CHECKBOX_TRUE_LIST else False


def parse_multiple_select(cell_value):
    cell_value = str(cell_value)
    values = cell_value.split('，') if '，' in cell_value else cell_value.split(',')
    options = []
    for value in values:
        option = value.strip(' ')
        if not option:
            continue
        options.append(option)
    return options


def parse_image(cell_value):
    cell_value = str(cell_value)
    return cell_value.split(' ')


def parse_number(cell_value):
    try:
        return float(cell_value)
    except:
        return ''

def parse_duration(cell_value):
    parsed_value = parse_number(cell_value)
    if not parsed_value:
        try:
            value_list = re.split('[:：]', str(cell_value))
            hours, minutes, seconds = value_list
            return 3600 * int(hours) + 60 * int(minutes) + int(seconds)
        except Exception as e:
            return ''
    return parsed_value



def parse_long_text(cell_value):
    cell_value = str(cell_value)
    checked_count = cell_value.count('[x]')
    unchecked_count = cell_value.count('[ ]')
    total = checked_count + unchecked_count

    href_reg = re.compile(HREF_REG)
    preview = href_reg.sub(' ', cell_value)
    preview = preview[:150].replace('\n', ' ')

    images = []
    links = []
    href_list = href_reg.findall(cell_value)
    for href in href_list:
        if re.search(LINK_REG_1, href):
            links.append(re.search(LINK_REG_1, href).group(1))
        elif re.search(LINK_REG_2, href):
            links.append(re.search(LINK_REG_2, href).group(1))
        elif re.search(IMAGE_REG_1, href):
            images.append(re.search(IMAGE_REG_1, href).group(1))
        elif re.search(IMAGE_REG_2, href):
            images.append(re.search(IMAGE_REG_2, href).group(1))

    return {
        'text': cell_value,
        'preview': preview,
        'checklist': {'completed': checked_count, 'total': total},
        'images': images,
        'links': links,
    }


def parse_excel_rows(sheet_rows, columns, head_index, max_column):
    """
    parse excel according to excel
    """
    from dtable_events.dtable_io import dtable_io_logger

    if head_index < 0:
        value_rows = sheet_rows
    else:
        value_rows = sheet_rows[head_index + 1:]
    rows = []
    for row in value_rows:
        row_data = {}
        for index in range(max_column):
            if not columns[index]:
                continue
            try:
                cell_value = get_excel_cell_value(row, index)
                column_name = columns[index]['name']
                column_type = columns[index]['type']
                if cell_value is None:
                    continue
                if isinstance(cell_value, datetime) or isinstance(cell_value, time):  # JSON serializable
                    cell_value = str(cell_value)

                if column_type == 'number':
                    row_data[column_name] = cell_value
                elif column_type == 'date':
                    row_data[column_name] = str(cell_value)
                elif column_type == 'long-text':
                    row_data[column_name] = parse_long_text(cell_value)
                elif column_type == 'checkbox':
                    row_data[column_name] = parse_checkbox(cell_value)
                elif column_type == 'multiple-select':
                    row_data[column_name] = parse_multiple_select(cell_value)
                else:
                    row_data[column_name] = str(cell_value)
            except Exception as e:
                dtable_io_logger.exception(e)
        if row_data:
            rows.append(row_data)

    return rows


def guess_column_type(value_list):
    from dtable_events.dtable_io import dtable_io_logger

    try:
        type_list = []
        for cell_value in value_list:
            # cell_value may be zero
            if cell_value is None or cell_value == '':
                continue
            elif isinstance(cell_value, int) or isinstance(cell_value, float):
                column_type = 'number'
            elif isinstance(cell_value, datetime):
                column_type = 'date'
            elif isinstance(cell_value, time):
                column_type = 'text'
            elif '\n' in cell_value:
                column_type = 'long-text'
            elif cell_value in CHECKBOX_STRING_LIST:
                column_type = 'checkbox'
            elif (',' in cell_value or '，' in cell_value) \
                    and ('{' not in cell_value):
                column_type = 'multiple-select'
                multiple_value = cell_value.split('，') if '，' in cell_value else cell_value.split(',')
                for value in multiple_value:
                    if len(value.strip(' ')) > 20:
                        # more than 20 characters.
                        column_type = 'text'
            else:
                column_type = 'text'
            type_list.append(column_type)

        max_column_type = max(type_list, key=type_list.count) if type_list else None

        if max_column_type == 'number' and len(set(type_list)) != 1:
            max_column_type = 'text'

        column_data = None
        if max_column_type == 'multiple-select':
            multiple_list = []
            for cell_value in value_list:
                if cell_value is None:
                    continue
                cell_value = str(cell_value)
                multiple_value = cell_value.split('，') if '，' in cell_value else cell_value.split(',')
                for value in multiple_value:
                    value = value.strip(' ')
                    if value not in multiple_list:
                        multiple_list.append(value)
            column_data = {'options': [{'name': value} for value in multiple_list]}

        return max_column_type, column_data
    except Exception as e:
        dtable_io_logger.exception(e)
        return 'text', None


def get_excel_cell_value(row, index):
    try:
        return row[index].value
    except:
        return None


def parse_excel_columns(sheet_rows, head_index, max_column):
    if head_index == -1:
        empty_cell = EmptyCell()
        head_row = [empty_cell] * max_column
        value_rows = sheet_rows
    else:
        head_row = sheet_rows[head_index]
        value_rows = sheet_rows[head_index + 1:]
    columns = []
    column_name_set = set()

    for index in range(max_column):
        name = get_excel_cell_value(head_row, index)
        # remove whitespace from both ends of name and BOM char(\ufeff)
        column_name = str(name).replace('\ufeff', '').strip() if name else 'Field' + str(index + 1)

        column_name = column_name.replace('`', '_').replace('{', '_').replace('}', '_').replace('.', '_')
        column_name = get_non_duplicated_name(column_name, column_name_set)
        column_name_set.add(column_name)

        value_list = [get_excel_cell_value(row, index) for row in value_rows[:200]]
        # Check the first 200 rows of data to determine column type
        column_type, column_data = guess_column_type(value_list)

        if name and not column_type:
            column_type = 'text'

        if not column_type:
            column = {}
        else:
            column = {
                'name': column_name,
                'type': column_type,
                'data': column_data
            }
        columns.append(column)

    return columns


def parse_excel(file_path, exist_tables=None):
    from dtable_events.dtable_io import dtable_io_logger
    if exist_tables is None:
        exist_tables = []

    tables = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    for sheet in wb:
        try:
            sheet_rows = list(sheet.rows)
        except Exception as e:
            raise ExcelFormatError
        if not sheet_rows:
            continue

        table_name = sheet.title
        table_name = table_name.replace('`', '_').replace('\\', '_').replace('/', '_')
        table_name = get_non_duplicated_name(table_name, exist_tables)

        # the sheet has some rows, but sheet.max_row maybe get None
        max_row = sheet.max_row if isinstance(sheet.max_row, int) else len(sheet_rows)
        max_column = sheet.max_column if isinstance(sheet.max_column, int) else len(sheet_rows[0])
        if not max_row or not max_column:
            continue

        dtable_io_logger.info(
            'parse sheet: %s, rows: %d, columns: %d' % (table_name, max_row, max_row))

        if max_row > 50000:
            max_row = 50000  # rows limit
        if max_column > 500:
            max_column = 500  # columns limit

        sheet_rows = sheet_rows[:max_row]
        head_index = 0
        columns = parse_excel_columns(sheet_rows, head_index, max_column)
        rows = parse_excel_rows(sheet_rows, columns, head_index, max_column)
        new_columns = [column for column in columns if column]

        dtable_io_logger.info(
            'got table: %s, rows: %d, columns: %d' % (table_name, len(rows), len(new_columns)))

        table = {
            'name': table_name,
            'rows': rows,
            'columns': new_columns,
            'head_index': head_index,
            'max_row': max_row,
            'max_column': len(new_columns),
        }
        tables.append(table)
    wb.close()
    if not tables:
        table = {
            'name': list(wb) and list(wb)[0].title or 'sheet1',
            'rows': [],
            'columns': [],
            'max_row': 0,
            'max_column': 0,
        }
        tables.append(table)

    return json.dumps(tables)


def parse_dtable_csv_columns(sheet_rows, max_column):
    head_row = sheet_rows[0]
    value_rows = sheet_rows[1:]

    columns = []
    column_name_set = set()
    for index in range(max_column):
        name = get_csv_cell_value(head_row, index)
        column_name = str(name).replace('\ufeff', '').strip() if name else 'Field' + str(index + 1)

        column_name = column_name.replace('`', '_').replace('{', '_').replace('}', '_').replace('.', '_')
        column_name = get_non_duplicated_name(column_name, column_name_set)
        column_name_set.add(column_name)

        value_list = [get_csv_cell_value(row, index) for row in value_rows[:200]]
        # Check the first 200 rows of data to determine column type
        column_type, column_data = guess_column_type(value_list)

        if name and not column_type:
            column_type = 'text'

        if not column_type:
            column = {}
        else:
            column = {
                'name': column_name,
                'type': column_type,
                'data': column_data
            }
        columns.append(column)

    return columns


def get_csv_cell_value(row, index):
    try:
        return row[index].strip()
    except:
        return None


def parse_dtable_csv_rows(sheet_rows, columns, max_column):
    from dtable_events.dtable_io import dtable_io_logger

    value_rows = sheet_rows[1:]
    rows = []
    for row in value_rows:
        row_data = {}
        for index in range(max_column):
            if not columns[index]:
                continue
            try:
                cell_value = get_csv_cell_value(row, index)
                column_name = columns[index]['name']
                column_type = columns[index]['type']
                if cell_value is None:
                    continue
                if column_type == 'long-text':
                    row_data[column_name] = parse_long_text(cell_value)
                elif column_type == 'checkbox':
                    row_data[column_name] = parse_checkbox(cell_value)
                elif column_type == 'multiple-select':
                    row_data[column_name] = parse_multiple_select(cell_value)
                else:
                    row_data[column_name] = str(cell_value)
            except Exception as e:
                dtable_io_logger.exception(e)
        if row_data:
            rows.append(row_data)

    return rows


def parse_dtable_csv(file_path, dtable_name, exist_tables=None):
    from dtable_events.dtable_io import dtable_io_logger
    if exist_tables is None:
        exist_tables = []

    dtable_name = dtable_name.replace('`', '_').replace('\\', '_').replace('/', '_')
    dtable_name = get_non_duplicated_name(dtable_name, exist_tables)

    csv_file = get_csv_file(file_path)
    csv_file = StringIO(csv_file)

    tables = []
    delimiter = guess_delimiter(deepcopy(csv_file))
    csv_rows = [row for row in csv.reader(csv_file, delimiter=delimiter)]

    if not csv_rows:
        table = {
            'name': dtable_name,
            'rows': [],
            'columns': [],
            'max_row': 0,
            'max_column': 0,
        }
        tables.append(table)
        return json.dumps(tables)

    csv_head = csv_rows[0]
    max_row = len(csv_rows)
    max_column = len(csv_head)

    if max_row > 50000:
        max_row = 50000
    if max_column > 500:
        max_column = 500

    csv_rows = csv_rows[:max_row]
    columns = parse_dtable_csv_columns(csv_rows, max_column)
    rows = parse_dtable_csv_rows(csv_rows, columns, max_column)

    new_columns = [column for column in columns if column]

    dtable_io_logger.info(
        'got table: %s, rows: %d, columns: %d' % (dtable_name, len(rows), len(new_columns)))

    table = {
        'name': dtable_name,
        'rows': rows,
        'columns': new_columns,
        'max_row': max_row,
        'max_column': len(new_columns),
    }
    tables.append(table)
    return json.dumps(tables)


def parse_and_import_excel_csv_to_dtable(repo_id, dtable_name, dtable_uuid, username, file_type, lang):
    base_path = os.path.join(EXCEL_IMPORT_DIR, repo_id)
    try:
        if file_type == 'xlsx':
            tmp_file_path = os.path.join(base_path, dtable_name + '.xlsx')
            content = parse_excel(tmp_file_path)
        else:
            tmp_file_path = os.path.join(base_path, dtable_name + '.csv')
            content = parse_dtable_csv(tmp_file_path, dtable_name)
    finally:
        # delete excel、csv  file
        clear_tmp_file(tmp_file_path)

    # import json file to dtable-server
    upload_excel_json_to_dtable_server(username, dtable_uuid, content, lang)


def parse_and_import_excel_csv_to_table(file_name, dtable_uuid, username, file_type, lang):
    base_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid)
    exist_table_names = get_table_names_by_dtable_server(username, dtable_uuid)
    try:
        # file_type is xlsx or csv
        if file_type == 'xlsx':
            tmp_file_path = os.path.join(base_path, file_name + '.xlsx')
            content = parse_excel(tmp_file_path, exist_tables=exist_table_names)
        else:
            tmp_file_path = os.path.join(base_path, file_name + '.csv')
            content = parse_dtable_csv(tmp_file_path, file_name, exist_tables=exist_table_names)
    finally:
        # delete excel or csv file
        clear_tmp_file(tmp_file_path)
    # import json file to dtable-server
    upload_excel_json_add_table_to_dtable_server(username, dtable_uuid, content, lang)


def parse_and_update_file_to_table(file_name, username, dtable_uuid, table_name, selected_columns, file_type):
    related_users = get_related_nicknames_from_dtable(dtable_uuid, username, 'r')
    name_to_email = {user.get('name'): user.get('email') for user in related_users}

    dtable_server_url = get_inner_dtable_server_url()
    dtable_server_api = DTableServerAPI(username, dtable_uuid, dtable_server_url)
    columns = dtable_server_api.list_columns(table_name)

    try:
        # file_type is xlsx or csv
        if file_type == 'xlsx':
            tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.xlsx')
            file_rows = parse_dtable_excel_file(tmp_file_path, table_name, columns, name_to_email)
        else:
            tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.csv')
            file_rows = parse_csv_file(tmp_file_path, file_name, table_name, columns, name_to_email)
    finally:
        clear_tmp_file(tmp_file_path)

    file_rows = file_rows[0].get('rows', [])
    key_columns = selected_columns.split(',')

    dtable_db_api = DTableDBAPI(username, dtable_uuid, INNER_DTABLE_DB_URL)
    dtable_rows = get_rows_from_dtable_db(dtable_db_api, table_name)

    dtable_col_name_to_column = {col['name']: col for col in columns}

    insert_rows, update_rows, excel_select_column_options = \
        get_insert_update_rows(dtable_col_name_to_column, file_rows, dtable_rows, key_columns, need_select_option=True)

    # add select column options
    for col_name, excel_options in excel_select_column_options.items():
        column = dtable_col_name_to_column.get(col_name)
        dtable_options = column.get('data').get('options')
        to_be_added_options = excel_options - set([op.get('name') for op in dtable_options])
        if to_be_added_options:
            options = [gen_random_option(option) for option in to_be_added_options]
            dtable_server_api.add_column_options(table_name, col_name, options)

    update_rows_by_dtable_db(dtable_db_api, update_rows, table_name)
    append_rows_by_dtable_server(dtable_server_api, insert_rows, table_name)


def parse_and_append_excel_csv_to_table(username, file_name, dtable_uuid, table_name, file_type):
    related_users = get_related_nicknames_from_dtable(dtable_uuid, username, 'r')
    name_to_email = {user.get('name'): user.get('email') for user in related_users}

    dtable_server_url = get_inner_dtable_server_url()
    dtable_server_api = DTableServerAPI(username, dtable_uuid, dtable_server_url)
    columns = dtable_server_api.list_columns(table_name)

    try:
        # file_type is xlsx or csv
        if file_type == 'xlsx':
            tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.xlsx')
            content = parse_dtable_excel_file(tmp_file_path, table_name, columns, name_to_email)
        else:
            tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.csv')
            content = parse_csv_file(tmp_file_path, file_name, table_name, columns, name_to_email)
    finally:
        # delete excel or csv
        clear_tmp_file(tmp_file_path)

    rows = content[0]['rows']
    dtable_col_name_to_column = {col['name']: col for col in columns}
    excel_select_column_options = extract_select_options(rows, dtable_col_name_to_column)

    # add single-select or multiple-select column options
    for col_name, excel_options in excel_select_column_options.items():
        column = dtable_col_name_to_column.get(col_name)
        dtable_options = column.get('data').get('options')
        to_be_added_options = excel_options - set([op.get('name') for op in dtable_options])
        if to_be_added_options:
            options = [gen_random_option(option) for option in to_be_added_options]
            dtable_server_api.add_column_options(table_name, col_name, options)

    append_rows_by_dtable_server(dtable_server_api, rows, table_name)


def parse_excel_csv_to_json(username, repo_id, file_name, file_type, parse_type, dtable_uuid):
    exist_table_names = []
    if parse_type == 'dtable':
        base_path = os.path.join(EXCEL_IMPORT_DIR, repo_id)
    else:
        base_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid)
        exist_table_names = get_table_names_by_dtable_server(username, dtable_uuid)

    try:
        # file_type is xlsx or csv
        if file_type == 'xlsx':
            tmp_file_path = os.path.join(base_path, file_name + '.xlsx')
            content = parse_excel(tmp_file_path, exist_tables=exist_table_names)
        else:
            tmp_file_path = os.path.join(base_path, file_name + '.csv')
            content = parse_dtable_csv(tmp_file_path, file_name, exist_tables=exist_table_names)
    finally:
        clear_tmp_file(tmp_file_path)

    # save tmp json file
    temp_json_path = os.path.join(base_path, file_name + '.json')
    save_file_by_path(temp_json_path, content)


def import_excel_csv_by_dtable_server(username, repo_id, dtable_uuid, dtable_name, included_tables, lang):
    # get json file
    tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, repo_id, dtable_name + '.json')
    with open(tmp_file_path, 'r') as f:
        json_file = f.read()

    clear_tmp_file(tmp_file_path)

    tables = [table for table in json.loads(json_file) if table.get('name') in included_tables]
    if not tables:
        raise Exception('tables invalid.')
    json_file = json.dumps(tables)

    # upload json file to dtable-server
    upload_excel_json_to_dtable_server(username, dtable_uuid, json_file, lang)


def import_excel_csv_add_table_by_dtable_server(username, dtable_uuid, dtable_name, included_tables, lang):
    # get json file
    tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, dtable_name + '.json')
    with open(tmp_file_path, 'r') as f:
        json_file = f.read()

    # delete tmp json file
    clear_tmp_file(tmp_file_path)

    tables = [table for table in json.loads(json_file) if table.get('name') in included_tables]
    if not tables:
        raise Exception('tables invalid.')
    json_file = json.dumps(tables)

    # upload json file to dtable-server
    upload_excel_json_add_table_to_dtable_server(username, dtable_uuid, json_file, lang)


def append_parsed_file_by_dtable_server(username, dtable_uuid, file_name, table_name):
    # get json file
    tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.json')
    with open(tmp_file_path, 'r') as f:
        json_file = f.read()

    # # delete json file
    clear_tmp_file(tmp_file_path)

    # upload json file to dtable-server
    rows = json.loads(json_file)[0]['rows']

    dtable_server_url = get_inner_dtable_server_url()
    dtable_server_api = DTableServerAPI(username, dtable_uuid, dtable_server_url)
    columns = dtable_server_api.list_columns(table_name)

    dtable_col_name_to_column = {col['name']: col for col in columns}
    excel_select_column_options = extract_select_options(rows, dtable_col_name_to_column)

    # add single-select or multiple-select column options
    for col_name, excel_options in excel_select_column_options.items():
        column = dtable_col_name_to_column.get(col_name)
        # data is None if table is new
        dtable_options = []
        if column.get('data') and column.get('data').get('options'):
            dtable_options = column.get('data').get('options')
        to_be_added_options = excel_options - set([op.get('name') for op in dtable_options])

        if to_be_added_options:
            options = [gen_random_option(option) for option in to_be_added_options]
            dtable_server_api.add_column_options(table_name, col_name, options)

    append_rows_by_dtable_server(dtable_server_api, rows, table_name)


def parse_append_excel_csv_upload_file_to_json(file_name, username, dtable_uuid, table_name, file_type):
    # parse
    related_users = get_related_nicknames_from_dtable(dtable_uuid, username, 'r')
    name_to_email = {user.get('name'): user.get('email') for user in related_users}

    dtable_server_url = get_inner_dtable_server_url()
    dtable_server_api = DTableServerAPI(username, dtable_uuid, dtable_server_url)
    columns = dtable_server_api.list_columns(table_name)
    if file_type == 'csv':
        tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.csv')
        tables = parse_csv_file(tmp_file_path, file_name, table_name, columns, name_to_email)
    else:
        tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.xlsx')
        tables = parse_dtable_excel_file(tmp_file_path, table_name, columns, name_to_email)

    # clear excel or csv
    clear_tmp_file(tmp_file_path)

    # save tmp json file
    temp_json_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.json')
    save_file_by_path(temp_json_path, json.dumps(tables))


def get_update_row_data(excel_row, dtable_row, excel_col_name_to_type):
    update_excel_row = {}
    for col_name in excel_col_name_to_type:
        excel_cell_val = excel_row.get(col_name, '')
        dtable_cell_val = dtable_row.get(col_name, '')
        column_type = excel_col_name_to_type.get(col_name)
        if column_type == 'multiple-select' or column_type == 'collaborator':
            if not dtable_cell_val:
                dtable_cell_val = []
            if not excel_cell_val:
                excel_cell_val = []
            excel_cell_val.sort()
            dtable_cell_val.sort()
        elif column_type == 'date' and excel_cell_val and dtable_cell_val:
            # dtable row value like 2021-12-03 00:00 or 2021-12-03, excel row like 2021-12-03 00:00:00
            excel_cell_val = excel_cell_val[0:len(dtable_cell_val)]
        elif column_type == 'checkbox' and not excel_cell_val:
            excel_cell_val = False
        dtable_cell_val = '' if dtable_cell_val is None else dtable_cell_val
        excel_cell_val = '' if excel_cell_val is None else excel_cell_val

        if excel_cell_val != dtable_cell_val:
            update_excel_row[col_name] = excel_cell_val
    if update_excel_row:
        return {'row_id': dtable_row.get('_id'), 'row': update_excel_row}


def get_dtable_row_data(dtable_rows, key_columns, excel_col_name_to_type):
    dtable_row_data = {}
    for row in dtable_rows:
        key = str(hash('-'.join([str(get_cell_value(row, col, excel_col_name_to_type)) for col in key_columns])))
        if dtable_row_data.get(key):
            # only deal first row
            continue
        else:
            dtable_row_data[key] = row
    return dtable_row_data


def update_parsed_file_by_dtable_server(username, dtable_uuid, file_name, table_name, selected_columns):
    # get json file
    tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.json')
    with open(tmp_file_path, 'r') as f:
        json_file = f.read()

    # delete tmp json file
    clear_tmp_file(tmp_file_path)

    excel_rows = json.loads(json_file)
    excel_rows = excel_rows[0].get('rows', [])
    key_columns = selected_columns.split(',')

    dtable_db_api = DTableDBAPI(username, dtable_uuid, INNER_DTABLE_DB_URL)
    dtable_rows = get_rows_from_dtable_db(dtable_db_api, table_name)

    dtable_server_url = get_inner_dtable_server_url()
    dtable_server_api = DTableServerAPI(username, dtable_uuid, dtable_server_url)

    columns = dtable_server_api.list_columns(table_name)

    dtable_col_name_to_column = {col['name']: col for col in columns}

    insert_rows, update_rows, excel_select_column_options = \
        get_insert_update_rows(dtable_col_name_to_column, excel_rows, dtable_rows, key_columns, need_select_option=True)

    # add select column options
    for col_name, excel_options in excel_select_column_options.items():
        column = dtable_col_name_to_column.get(col_name)
        col_data = column.get('data')
        dtable_options = col_data and col_data.get('options') or []
        to_be_added_options = excel_options - set([op.get('name') for op in dtable_options])
        if to_be_added_options:
            options = [gen_random_option(option) for option in to_be_added_options]
            dtable_server_api.add_column_options(table_name, col_name, options)

    update_rows_by_dtable_db(dtable_db_api, update_rows, table_name)
    append_rows_by_dtable_server(dtable_server_api, insert_rows, table_name)


def get_cell_value(row, col, excel_col_name_to_type):
    cell_value = row.get(col)
    col_type = excel_col_name_to_type.get(col)
    if col_type == 'number':
        if isinstance(cell_value, float):
            cell_value = str(cell_value).rstrip('0')
            cell_value = int(cell_value.rstrip('.')) if cell_value.endswith('.') else float(cell_value)

    cell_value = '' if cell_value is None else cell_value
    return cell_value


def get_insert_update_rows(dtable_col_name_to_column, excel_rows, dtable_rows, key_columns, need_select_option=False):
    if not excel_rows:
        return [], [], {}
    update_rows = []
    insert_rows = []
    excel_select_column_options = {}
    excel_col_name_to_type = {col_name: dtable_col_name_to_column.get(col_name).get('type') for col_name in excel_rows[0].keys()
                              if dtable_col_name_to_column.get(col_name, {}).get('type') in UPDATE_TYPE_LIST}

    dtable_row_data = get_dtable_row_data(dtable_rows, key_columns, excel_col_name_to_type)
    keys_of_excel_rows = {}
    for excel_row in excel_rows:
        excel_row = {col_name: excel_row.get(col_name) for col_name in excel_row if excel_col_name_to_type.get(col_name)}
        key = str(hash('-'.join([str(get_cell_value(excel_row, col, excel_col_name_to_type)) for col in key_columns])))
        if keys_of_excel_rows.get(key):
            continue
        keys_of_excel_rows[key] = True

        if need_select_option:
            # get column options for update single-select or multiple-select columns
            for col_name in excel_row:
                col_type = excel_col_name_to_type.get(col_name)
                cell_value = excel_row.get(col_name)
                if not cell_value:
                    continue
                if col_type in ['multiple-select', 'single-select']:
                    col_options = excel_select_column_options.get(col_name, set())
                    if not col_options:
                        excel_select_column_options[col_name] = col_options
                    if col_type == 'multiple-select':
                        col_options.update(set(cell_value))
                    else:
                        col_options.add(cell_value)

        dtable_row = dtable_row_data.get(key)
        if not dtable_row:
            insert_rows.append(excel_row)
        else:
            update_row = get_update_row_data(excel_row, dtable_row, excel_col_name_to_type)
            if update_row:
                update_rows.append(update_row)
    return insert_rows, update_rows, excel_select_column_options


def parse_dtable_excel_file(file_path, table_name, columns, name_to_email):
    from dtable_events.dtable_io import dtable_io_logger

    tables = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])

    try:
        sheet_rows = list(sheet.rows)
    except Exception as e:
        raise ExcelFormatError
    if not sheet_rows:
        wb.close()
        table = {
            'name': table_name,
            'rows': [],
            'columns': columns,
            'max_row': 0,
            'max_column': 0,
        }
        tables.append(table)
        return tables

    # the sheet has some rows, but sheet.max_row maybe get None
    max_row = sheet.max_row if isinstance(sheet.max_row, int) else len(sheet_rows)
    max_column = sheet.max_column if isinstance(sheet.max_column, int) else len(sheet_rows[0])

    dtable_io_logger.info(
        'parse sheet: %s, rows: %d, columns: %d' % (sheet.title, max_row, max_column))

    if max_row > 50000:
        max_row = 50000  # rows limit
    if max_column > 500:
        max_column = 500  # columns limit

    if max_column > len(columns):
        max_column = len(columns)

    sheet_rows = sheet_rows[:max_row]
    rows = parse_dtable_excel_rows(sheet_rows, columns, len(columns), name_to_email)

    dtable_io_logger.info(
        'got table: %s, rows: %d, columns: %d' % (sheet.title, len(rows), max_column))

    table = {
        'name': table_name,
        'rows': rows,
        'columns': columns,
        'max_row': max_row,
        'max_column': max_column,
    }
    tables.append(table)
    wb.close()

    return tables


def parse_update_excel_upload_excel_to_json(file_name, username, dtable_uuid, table_name):
    related_users = get_related_nicknames_from_dtable(dtable_uuid, username, 'r')
    name_to_email = {user.get('name'): user.get('email') for user in related_users}

    dtable_server_url = get_inner_dtable_server_url()
    dtable_server_api = DTableServerAPI(username, dtable_uuid, dtable_server_url)
    columns = dtable_server_api.list_columns(table_name)

    tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.xlsx')
    content = parse_dtable_excel_file(tmp_file_path, table_name, columns, name_to_email)

    clear_tmp_file(tmp_file_path)

    tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.json')
    save_file_by_path(tmp_file_path, json.dumps(content))


def parse_dtable_excel_rows(sheet_rows, columns, column_length, name_to_email):
    """
    parse excel according to dtable
    """
    from dtable_events.dtable_io import dtable_io_logger
    from dtable_events.utils import get_location_tree_json

    value_rows = sheet_rows[1:]
    sheet_head = sheet_rows[0]
    head_dict = {sheet_head[index].value: index for index in range(len(sheet_head))}
    rows = []

    location_tree = get_location_tree_json()

    for row in value_rows:
        row_data = {}
        for index in range(column_length):
            column_name = columns[index]['name']
            if head_dict.get(column_name) is None:
                continue
            row_index = head_dict.get(column_name)
            try:
                cell_value = get_excel_cell_value(row, row_index)
                column_type = columns[index]['type']
                if cell_value is None:
                    row_data[column_name] = None
                    continue
                row_data[column_name] = parse_row(column_type, cell_value, name_to_email, location_tree=location_tree)
            except Exception as e:
                dtable_io_logger.exception(e)
                row_data[column_name] = None
        if row_data:
            rows.append(row_data)
    return rows


def parse_csv_file(file_path, file_name, table_name, columns, name_to_email):
    from dtable_events.dtable_io import dtable_io_logger

    csv_file = get_csv_file(file_path)

    tables = []
    max_column = 500  # columns limit
    rows, max_column, csv_row_num, csv_column_num = parse_csv_rows(StringIO(csv_file), columns, max_column, name_to_email)
    dtable_io_logger.info(
        'parse csv: %s, rows: %d, columns: %d' % (file_name, csv_row_num, csv_column_num))

    max_row = csv_row_num
    if csv_row_num > 50000:
        max_row = 50000  # rows limit
    rows = rows[:max_row]

    dtable_io_logger.info(
        'got table: %s, rows: %d, columns: %d' % (file_name, len(rows), max_column))

    table = {
        'name': table_name,
        'rows': rows,
        'columns': columns,
        'max_row': max_row,
        'max_column': max_column,
    }
    tables.append(table)
    return tables


def parse_update_csv_upload_csv_to_json(file_name, username, dtable_uuid, table_name):
    related_users = get_related_nicknames_from_dtable(dtable_uuid, username, 'r')
    name_to_email = {user.get('name'): user.get('email') for user in related_users}

    dtable_server_url = get_inner_dtable_server_url()
    dtable_server_api = DTableServerAPI(username, dtable_uuid, dtable_server_url)
    columns = dtable_server_api.list_columns(table_name)

    tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.csv')
    content = parse_csv_file(tmp_file_path, file_name, table_name, columns, name_to_email)

    clear_tmp_file(tmp_file_path)

    tmp_file_path = os.path.join(EXCEL_IMPORT_DIR, dtable_uuid, file_name + '.json')
    save_file_by_path(tmp_file_path, json.dumps(content))


def guess_delimiter(csv_file):
    line = csv_file.readline()

    if not line:
        return ','
    comma_count = line.count(',')
    semicolon_count = line.count(';')
    delimiter = comma_count >= semicolon_count and ',' or ';'

    return delimiter


def parse_csv_rows(csv_file, columns, max_column, name_to_email):
    from dtable_events.dtable_io import dtable_io_logger
    from dtable_events.utils import get_location_tree_json

    rows = []
    delimiter = guess_delimiter(deepcopy(csv_file))
    csv_rows = [row for row in csv.reader(csv_file, delimiter=delimiter)]

    if not csv_rows:
        return rows, 0, 0, 0

    location_tree = get_location_tree_json()

    csv_head = csv_rows[0]
    csv_column_num = len(csv_head)
    table_column_num = len(columns)
    if csv_column_num < max_column:
        max_column = csv_column_num
    if table_column_num > csv_column_num:
        max_column = table_column_num

    csv_head_dict = {csv_head[index].strip(): index for index in range(csv_column_num)}
    csv_row_num = 0
    for csv_row in csv_rows[1:]:
        csv_row_num += 1
        row_data = {}
        for index in range(table_column_num):
            column_name = columns[index]['name']
            row_index = csv_head_dict.get(column_name)
            if row_index is None:
                continue
            try:
                cell_value = get_csv_cell_value(csv_row, row_index)
                column_type = columns[index]['type']
                if cell_value is None:
                    row_data[column_name] = None
                    continue
                parsed_value = parse_row(column_type, cell_value, name_to_email, location_tree=location_tree)
                row_data[column_name] = parsed_value
            except Exception as e:
                dtable_io_logger.exception(e)
                row_data[column_name] = None
        if row_data:
            rows.append(row_data)
    return rows, max_column, csv_row_num, csv_column_num


def parse_row(column_type, cell_value, name_to_email, location_tree=None):
    if isinstance(cell_value, datetime):  # JSON serializable
        cell_value = str(cell_value)
    if isinstance(cell_value, str):
        cell_value = cell_value.strip()
    if column_type in ('number', 'rate'):
        return parse_number(cell_value)
    elif column_type == 'duration':
        return parse_duration(cell_value)
    elif column_type == 'date':
        return str(cell_value)
    elif column_type == 'long-text':
        return parse_long_text(cell_value)
    elif column_type == 'checkbox':
        return parse_checkbox(cell_value)
    elif column_type == 'multiple-select':
        return parse_multiple_select(cell_value)
    elif column_type in ('url', 'email'):
        return str(cell_value)
    elif column_type == 'text':
        return str(cell_value)
    elif column_type == 'file':
        return None
    elif column_type == 'image':
        return parse_image(cell_value)
    elif column_type == 'single-select':
        return str(cell_value)
    elif column_type == 'link':
        return None
    elif column_type == 'button':
        return None
    elif column_type == 'geolocation':
        return parse_geolocation_from_tree(location_tree, cell_value)
    elif column_type in ('creator', 'last-modifier', 'ctime', 'mtime', 'formula', 'link-formula', 'auto-number'):
        return None
    elif column_type == 'collaborator':
        cell_value = parse_collaborator(cell_value, name_to_email)
        return cell_value
    else:
        return str(cell_value)


def parse_collaborator(cell_value, name_to_email):
    if not isinstance(cell_value, str):
        return []
    users = re.split('[,，]', cell_value)
    email_list = []
    for user in users:
        email = name_to_email.get(user.strip())
        if email:
            email_list.append(email)
    return email_list


def get_summary(summary, summary_col_info, column_name, head_name_to_head):
    summary_type = summary_col_info.get(column_name, 'sum').lower()
    column_info = head_name_to_head.get(column_name)
    # return summary info if column type is formula and result type is number for excel value
    # because grouped summary row does not contain format symbol like $, ￥, %, etc
    if column_info and column_info.get('type') == 'formula' and column_info.get('data').get('result_type') == 'number':
        return parse_summary_value(summary.get(summary_type), column_info.get('data'))
    if column_info and column_info.get('type') == 'duration':
        return format_duration(summary.get(summary_type), column_info.get('data'))
    return summary.get(summary_type)


def parse_geolocation(cell_data):
    if not isinstance(cell_data, dict):
        return str(cell_data)
    if 'country_region' in cell_data:
        return cell_data['country_region']
    elif 'lng' in cell_data:
        return str(cell_data['lng']) + ', ' + str(cell_data['lat'])
    elif 'province' in cell_data:
        value = str(cell_data['province'])
        if 'city' in cell_data:
            value = '%s%s' % (value, cell_data['city'])
        if 'district' in cell_data:
            value = '%s%s' % (value, cell_data['district'])
        if 'detail' in cell_data:
            value = '%s%s' % (value, cell_data['detail'])
        return value
    else:
        return str(cell_data)


def parse_link_formula(cell_data, email2nickname):
    from dtable_events.dtable_io import dtable_io_logger
    try:
        # collaborator
        if isinstance(cell_data, list) \
                and isinstance(cell_data[0], str) \
                and VIRTUAL_ID_EMAIL_DOMAIN in cell_data[0]:
            nickname_list = []
            for user in cell_data:
                nickname_list.append(email2nickname.get(user, ''))
            value = ', '.join(nickname_list)
        # ctime, mtime
        elif isinstance(cell_data, str) \
                and '+00:00' in cell_data \
                and 'T' in cell_data:
            utc_time = datetime.strptime(cell_data, '%Y-%m-%dT%H:%M:%S.%f+00:00')
            value = utc_to_tz(utc_time, timezone).strftime('%Y-%m-%d %H:%M:%S')
        # string
        else:
            value = cell_data2str(cell_data)
        return value
    except Exception as e:
        dtable_io_logger.warning(e)
        return cell_data2str(cell_data)


def cell_data2str(cell_data):
    if not cell_data:
        return ''
    elif isinstance(cell_data, list):
        return ' '.join(cell_data2str(item) for item in cell_data)
    else:
        return str(cell_data)


def parse_multiple_select_formula(cell_data):
    if isinstance(cell_data, list):
        return ', '.join(cell_data)
    else:
        return str(cell_data)


def convert_formula_number(value, column_data):
    decimal = column_data.get('decimal')
    thousands = column_data.get('thousands')
    precision = column_data.get('precision', 0)
    if decimal == 'comma':
        # decimal maybe dot or comma
        value = value.replace(',', '.')
    if thousands == 'space':
        # thousands maybe space, dot, comma or no
        value = value.replace(' ', '')
    elif thousands == 'dot':
        value = value.replace('.', '')
        if precision > 0 or decimal == 'dot':
            value = value[:-precision] + '.' + value[-precision:]
    elif thousands == 'comma':
        value = value.replace(',', '')

    return value


def parse_summary_value(cell_data, column_data):
    value = str(cell_data)
    precision = column_data.get('precision', 0)
    src_format = column_data.get('format')

    if src_format == 'percent':
        try:
            if is_int_str(value):
                value = str(int(value) * 100)
            else:
                value = str(float(value) * 100)
        except:
            pass
    elif src_format == 'duration':
        duration_format = column_data.get('duration_format', 'h:mm')
        duration_value = float(value)
        h_value = str(duration_value // 3600).split('.')[0]
        m_value = str((duration_value % 3600) // 60).split('.')[0]
        s_value = str(duration_value % 60).split('.')[0]
        if len(m_value) == 1:
            m_value = '0' + m_value
        if duration_format == 'h:mm':

            return h_value + ':' + m_value
        else:
            if len(s_value) == 1:
                s_value = '0' + s_value
            return h_value + ':' + m_value + ':' + s_value
    value_list = value.split('.')
    value_precision = len(value_list[1]) if (len(value_list) > 1) else 0

    if precision > 0 and precision > value_precision:
        if value_precision > 0:
            value = value + '0' * (precision - value_precision)
        else:
            value = value + '.' + '0' * (precision - value_precision)

    # add symbol
    if src_format == 'euro':
        value = '€' + value
    elif src_format == 'dollar':
        value = '$' + value
    elif src_format == 'yuan':
        value = '￥' + value
    elif src_format == 'percent':
        value = value + '%'
    elif src_format == 'custom_currency':
        currency_symbol = column_data.get('currency_symbol')
        currency_symbol_position = column_data.get('currency_symbol_position', 'before')
        if currency_symbol_position == 'before':
            value = currency_symbol + value
        else:
            value = value + currency_symbol
    return value


def parse_formula_number(cell_data, column_data, is_big_data_view=False):
    """
    parse formula number to regular format
    :param cell_data: value of cell (e.g. 1.25, ￥12.0, $10.20, €10.2, 0:02 or 10%, etc)
    :param column_data: info of formula column
    """
    src_format = column_data.get('format')
    value = str(cell_data)
    if src_format in ['euro', 'dollar', 'yuan']:
        value = is_big_data_view and value or value[1:]
    elif src_format == 'percent':
        value = is_big_data_view and value or value[:-1]
    elif src_format == 'custom_currency':
        currency_symbol = column_data.get('currency_symbol')
        currency_symbol_position = column_data.get('currency_symbol_position', 'before')
        if not is_big_data_view:
            if currency_symbol_position == 'before':
                value = value[len(currency_symbol):]
            else:
                value = value[:-len(currency_symbol)]
    value = convert_formula_number(value, column_data)
    number_format = '0'
    # src_format is None, if column format does not be changed
    if src_format == 'number' or not src_format:
        number_format = gen_decimal_format(value)
    elif src_format == 'percent' and isinstance(value, str):
        number_format = gen_decimal_format(value) + '%'
        try:
            value = float(value) / 100
        except Exception as e:
            pass
    elif src_format == 'euro':
        number_format = '"€"#,##' + gen_decimal_format(value)+'_-'
    elif src_format == 'dollar':
        number_format = '"$"#,##' + gen_decimal_format(value)+'_-'
    elif src_format == 'yuan':
        number_format = '"¥"#,##' + gen_decimal_format(value)+'_-'
    elif src_format == 'custom_currency':
        currency_symbol = column_data.get('currency_symbol')
        currency_symbol_position = column_data.get('currency_symbol_position', 'before')
        if currency_symbol_position == 'before':
            number_format = '"%s"#,##' % currency_symbol + gen_decimal_format(value) + '_-'
        else:
            number_format = gen_decimal_format(value) + currency_symbol
    try:
        if is_int_str(value):
            value = int(value)
        else:
            value = float(value)
    except Exception as e:
        pass
    return value, number_format


def convert_time_to_utc_str(time_str):
    if 'Z' in time_str:
        utc_time = datetime.strptime(time_str, '%Y-%m-%dT%H:%M:%S.%fZ')
    else:
        utc_time = datetime.strptime(time_str, '%Y-%m-%dT%H:%M:%S.%f+00:00')
    return utc_to_tz(utc_time, timezone).strftime('%Y-%m-%d %H:%M:%S')


def select_option_to_name(id2name, cell):
    if not cell.get('display_value'):
        return ''
    return id2name.get(cell.get('display_value'), '')

def email_to_nickname(email2nickname, cell):
    if not cell.get('display_value'):
        return ''
    return email2nickname.get(cell.get('display_value'), '')

def parse_link(column, cell_data, email2nickname):
    if isinstance(cell_data, list):
        if column.get('data').get('array_type') == ColumnTypes.SINGLE_SELECT:
            options = column.get('data').get('array_data', {}).get('options')
            id2name = {op.get('id'): op.get('name') for op in options}
            return ', '.join([select_option_to_name(id2name, cell) for cell in cell_data])
        elif column.get('data').get('array_type') in (ColumnTypes.CREATOR, ColumnTypes.LAST_MODIFIER):
            return ', '.join([email_to_nickname(email2nickname, cell) for cell in cell_data])
        elif column.get('data').get('array_type') in (ColumnTypes.CTIME, ColumnTypes.MTIME):
            return ', '.join([convert_time_to_utc_str(cell.get('display_value')) if cell.get('display_value') else '' for cell in cell_data])
        # display_value may be array
        return ', '.join([cell_data2str(cell.get('display_value')) if cell.get('display_value') else '' for cell in cell_data])
    else:
        return str(cell_data)


def is_int_str(num):
    return '.' not in str(num)


def gen_decimal_format(num):
    if is_int_str(num):
        return '0'

    decimal_cnt = len(str(num).split('.')[1])
    if decimal_cnt > 8:
        decimal_cnt = 8
    return '0.' + '0' * decimal_cnt


def check_and_replace_sheet_name(sheet_name):
    """/ ?\ * [ ] chars is invalid excel sheet name, replace these chars with _ """

    invalid_chars = ['/', '?', '\\', '*', '[', ']', ':']
    for char in invalid_chars:
        if char in sheet_name:
            sheet_name = sheet_name.replace(char, '_')
    return sheet_name


def add_nickname_to_cell(unknown_user_set, unknown_cell_list):

    unknown_user_id_list = list(unknown_user_set)
    step = 1000
    start = 0
    user_list = []
    for i in range(0, len(unknown_user_id_list), step):
        user_list += get_nicknames_from_dtable(unknown_user_id_list[start: start+step])
        start += step

    email2nickname = {nickname['email']: nickname['name'] for nickname in user_list}
    for excel_cell, user_info, col_type in unknown_cell_list:
        if col_type == ColumnTypes.COLLABORATOR:
            nickname_list, collaborator_email_list = user_info
            for email in collaborator_email_list:
                nickname_list.append(email2nickname.get(email, ''))
            excel_cell.value = ', '.join(nickname_list)
        else:
            excel_cell.value = email2nickname.get(user_info, '')


def parse_dtable_long_text(cell_value):
    if not isinstance(cell_value, str):
        return ''
    if cell_value.find('\n\n') == -1:
        return cell_value
    return parse_dtable_long_text(cell_value.replace('\n\n', '\n'))


def get_file_download_url(file_url, dtable_uuid, repo_id):
    from urllib.parse import unquote
    from seaserv import seafile_api
    from dtable_events.utils import uuid_str_to_36_chars, normalize_file_path, gen_file_get_url

    file_path = unquote('/'.join(file_url.split('/')[7:]).strip())
    asset_path = normalize_file_path(os.path.join('/asset', uuid_str_to_36_chars(dtable_uuid), file_path))
    asset_id = seafile_api.get_file_id_by_path(repo_id, asset_path)
    asset_name = os.path.basename(normalize_file_path(file_path))
    if not asset_id:
        return None

    token = seafile_api.get_fileserver_access_token(
        repo_id, asset_id, 'download', '', use_onetime=False
    )

    url = gen_file_get_url(token, asset_name)
    return url

def add_image_to_excel(ws, cell_value, col_num, row_num, dtable_uuid, repo_id, image_num, images_target_dir):
    import requests
    from openpyxl.drawing.image import Image
    from PIL import Image as PILImage
    from urllib.parse import unquote, urljoin, urlparse
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

    images = cell_value
    row_pos = str(row_num + 1)
    # set image cell height
    ws.row_dimensions[int(row_pos)].height = IMAGE_CELL_ROW_HEIGHT

    offset_increment = 0
    for image_url in images:
        if image_num >= EXPORT_IMAGE_LIMIT:
            return image_num
        real_image_url = urljoin(image_url, urlparse(image_url).path)

        image_name = unquote(real_image_url.split('/')[-1].strip())
        image_dir = os.path.join(images_target_dir, '/'.join(real_image_url.split('/')[7:-1]))
        os.makedirs(image_dir, exist_ok=True)

        image_download_url = get_file_download_url(real_image_url, dtable_uuid, repo_id)
        if not image_download_url:
            continue

        response = requests.get(image_download_url)
        image_content = response.content

        tmp_image_path = os.path.join(image_dir, image_name)
        with open(tmp_image_path, 'wb') as f:
            f.write(image_content)

        try:
            img = Image(tmp_image_path)
        except:
            continue
        image_format = img.format

        if image_format == 'mpo':
            continue

        # convert webp to png
        if image_format in ('webp', ):
            img = PILImage.open(tmp_image_path)
            img.load()
            image_name = image_name.split('.')[0] + '.png'
            new_tmp_image_path = os.path.join(image_dir, image_name)
            img.save(new_tmp_image_path, format='png')
            # remove webp image
            os.remove(tmp_image_path)
            img = Image(new_tmp_image_path)

        from_col_offset = FROM_COL_START_OFFSET + offset_increment
        from_row_offset = FROM_ROW_START_OFFSET + offset_increment
        to_col_offset = TO_COL_START_OFFSET + offset_increment
        to_row_offset = TO_ROW_START_OFFSET + offset_increment

        from_anchor = AnchorMarker(col_num, from_col_offset, row_num, from_row_offset)
        to_anchor = AnchorMarker(col_num + 1, to_col_offset, row_num + 1, to_row_offset)
        img.anchor = TwoCellAnchor('twoCell', from_anchor, to_anchor)

        ws.add_image(img)
        offset_increment += 20000
        image_num += 1
    return image_num


def format_time(cell_data):
    from dtable_events.dtable_io import dtable_io_logger

    try:
        timestamp = parser.isoparse(cell_data.strip()).timestamp()
        utc_time = datetime.utcfromtimestamp(timestamp)
        return utc_to_tz(utc_time, timezone)
    except Exception as e:
        dtable_io_logger.debug(e)
    return cell_data2str(cell_data)

def format_duration(cell_data, column_data):
    value = str(cell_data)
    duration_format = column_data.get('duration_format', 'h:mm')
    duration_value = float(value)
    h_value = str(duration_value // 3600).split('.')[0]
    m_value = str((duration_value % 3600) // 60).split('.')[0]
    s_value = str(duration_value % 60).split('.')[0]
    if len(m_value) == 1:
        m_value = '0' + m_value
    if duration_format == 'h:mm':
        return h_value + ':' + m_value
    else:
        if len(s_value) == 1:
            s_value = '0' + s_value
        return h_value + ':' + m_value + ':' + s_value


def handle_grouped_row(row, ws, cols_without_hidden, column_name_to_column, sub_level, summary_col_info, head_name_to_head, summaries):
    from openpyxl.cell import WriteOnlyCell
    cell_list = []
    is_first_column = True
    first_col_name = row.get('column_name')
    first_cell_value = row.get('cell_value')
    group_column = column_name_to_column.get(first_col_name)
    for column in cols_without_hidden:
        col_name = column.get('name')

        # parse group first column
        if is_first_column and not first_cell_value:
            c = WriteOnlyCell(ws, value=None)
        elif is_first_column:
            if group_column.get('type') == ColumnTypes.FORMULA and isinstance(group_column.get('data'), dict) \
                    and group_column.get('data').get('result_type') == 'number':

                first_cell_value = parse_summary_value(first_cell_value, group_column.get('data'))
                formula_value, number_format = parse_formula_number(first_cell_value, group_column.get('data'))
                c = WriteOnlyCell(ws, value=formula_value)
                c.number_format = number_format
            else:
                cell_value = cell_data2str(first_cell_value)
                c = WriteOnlyCell(ws, value=ILLEGAL_CHARACTERS_RE.sub('', cell_value))
        else:
            cell_value = summaries.get(col_name)
            if cell_value:
                # not empty means summary value
                # like {'price': {'sum': 89, 'average': 49.5, 'median': 49.5, 'max': 66, 'min': 233}, ...}
                cell_value = get_summary(cell_value, summary_col_info, col_name, head_name_to_head)
            c = WriteOnlyCell(ws, value=cell_value)

        try:
            c.fill = grouped_row_fills[sub_level]
        except:
            pass
        cell_list.append(c)
        is_first_column = False
    return cell_list


def handle_row(row, row_num, ws, email2nickname, unknown_user_set, unknown_cell_list, dtable_uuid, repo_id, image_param, cols_without_hidden, is_big_data_view=False):
    from openpyxl.cell import WriteOnlyCell
    cell_list = []
    col_num = 0
    for column in cols_without_hidden:
        col_name = column.get('name')
        col_type = column.get('type')
        cell_value = row.get(col_name)

        if not cell_value and not isinstance(cell_value, int) and not isinstance(cell_value, float):
            c = WriteOnlyCell(ws, value=None)

        # excel format see
        # https://support.office.com/en-us/article/Number-format-codes-5026bbd6-04bc-48cd-bf33-80f18b4eae68
        elif col_type == ColumnTypes.NUMBER:
            # if value cannot convert to float or int, just pass, e.g. empty srt ''
            try:
                if is_int_str(cell_value):
                    c = WriteOnlyCell(ws, value=int(cell_value))
                else:
                    c = WriteOnlyCell(ws, value=float(cell_value))
            except Exception as e:
                c = WriteOnlyCell(ws, value=None)
            else:
                column_data = column.get('data')
                if column_data:
                    formula_value, number_format = parse_formula_number(cell_value, column.get('data'), is_big_data_view)
                    c.number_format = number_format
                else:
                    c.number_format = gen_decimal_format(cell_value)

        elif col_type == ColumnTypes.DATE:
            c = WriteOnlyCell(ws, value=format_time(cell_value))
            if column.get('data'):
                c.number_format = column.get('data').get('format', '')
            else:
                c.number_format = 'YYYY-MM-DD'
        elif col_type in (ColumnTypes.CTIME, ColumnTypes.MTIME):
            c = WriteOnlyCell(ws, value=format_time(cell_value))
        elif col_type == ColumnTypes.DURATION:
            c = WriteOnlyCell(ws, value=format_duration(cell_value, column.get('data')))
        elif col_type == ColumnTypes.COLLABORATOR:
            nickname_list = []
            collaborator_email_list = []
            for user in cell_value:
                if not email2nickname.get(user, ''):
                    unknown_user_set.add(user)
                    collaborator_email_list.append(user)
                else:
                    nickname_list.append(email2nickname.get(user, ''))
            nicknames = ', '.join(nickname_list)
            c = WriteOnlyCell(ws, value=nicknames)
            if collaborator_email_list:
                unknown_cell_list.append((c, (nickname_list, collaborator_email_list), col_type))
            c.value = ', '.join(nickname_list)
        elif col_type == ColumnTypes.CREATOR:
            c = WriteOnlyCell(ws, value=email2nickname.get(cell_data2str(cell_value), ''))
            if not email2nickname.get(cell_data2str(cell_value), ''):
                unknown_user_set.add(cell_data2str(cell_value))
                unknown_cell_list.append((c, cell_data2str(cell_value), col_type))
        elif col_type == ColumnTypes.LAST_MODIFIER:
            c = WriteOnlyCell(ws, value=email2nickname.get(cell_data2str(cell_value), ''))
            if not email2nickname.get(cell_data2str(cell_value), ''):
                unknown_user_set.add(cell_data2str(cell_value))
                unknown_cell_list.append((c, cell_data2str(cell_value), col_type))
        elif col_type == ColumnTypes.FORMULA \
                and isinstance(column.get('data'), dict) and column.get('data').get('result_type') == 'number':
            formula_value, number_format = parse_formula_number(cell_value, column.get('data'), is_big_data_view)
            c = WriteOnlyCell(ws, value=formula_value)
            c.number_format = number_format
        elif col_type == ColumnTypes.IMAGE and cell_value and image_param['is_support']:
            c = WriteOnlyCell(ws)
            image_num = image_param.get('num')
            images_target_dir = image_param.get('images_target_dir')
            if image_num < EXPORT_IMAGE_LIMIT:
                num = add_image_to_excel(ws, cell_value, col_num, row_num, dtable_uuid, repo_id, image_num, images_target_dir)
                image_param['num'] = num
        else:
            if col_type == ColumnTypes.GEOLOCATION:
                cell_value = parse_geolocation(cell_value)
            elif col_type == ColumnTypes.LINK_FORMULA:
                cell_value = parse_link_formula(cell_value, email2nickname)
            elif col_type == ColumnTypes.MULTIPLE_SELECT:
                cell_value = parse_multiple_select_formula(cell_value)
            elif col_type == ColumnTypes.LINK:
                cell_value = parse_link(column, cell_value, email2nickname)
            elif col_type == ColumnTypes.LONG_TEXT:
                cell_value = parse_dtable_long_text(cell_value)
            else:
                cell_value = cell_data2str(cell_value)
            c = WriteOnlyCell(ws, value=ILLEGAL_CHARACTERS_RE.sub('', cell_value))

        cell_list.append(c)
        col_num += 1
    return cell_list


def write_xls_with_type(data_list, email2nickname, ws, row_num, dtable_uuid, repo_id, image_param, cols_without_hidden, column_name_to_column, is_group_view=False, summary_col_info=None, row_height='default', header_height='default', is_big_data_view=False):
    """ write listed data into excel
    """
    from dtable_events.dtable_io import dtable_io_logger
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.utils import get_column_letter
    from dtable_events.dtable_io.utils import width_transfer, height_transfer

    ws.row_dimensions[1].height = height_transfer(header_height) # set header height
    if row_num == 0:
        # write table head
        column_error_log_exists = False
        head_cell_list = []
        col_num = 0
        for col in cols_without_hidden:
            try:
                c = WriteOnlyCell(ws, value=col.get('name'))
                col_pos = get_column_letter(col_num + 1)
                if col.get('type') == ColumnTypes.IMAGE:

                    # set image column width
                    ws.column_dimensions[col_pos].width = IMAGE_CELL_COLUMN_WIDTH

                else:
                    col_width = col.get('width', 200)
                    col_width_xls = width_transfer(col_width)
                    ws.column_dimensions[col_pos].width = col_width_xls

            except Exception as e:
                if not column_error_log_exists:
                    dtable_io_logger.error('Error column in exporting excel: {}'.format(e))
                    column_error_log_exists = True
                c = WriteOnlyCell(ws, value=EXPORT2EXCEL_DEFAULT_STRING)
            head_cell_list.append(c)
            col_num += 1
        ws.append(head_cell_list)

    # write table data
    row_error_log_exists = False
    unknown_user_set = set()
    unknown_cell_list = []

    if is_group_view:
        row_list = []
        # for insert image
        row_num_info = {'row_num': row_num + 1}
        sub_level = 0
        handle_grouped_view_rows(data_list, row_num_info, ws, email2nickname, unknown_user_set, unknown_cell_list, dtable_uuid,
                         repo_id, image_param, cols_without_hidden, column_name_to_column, summary_col_info, row_list, sub_level)
    else:
        row_list = []
        for row in data_list:
            row_num += 1  # for big data view
            try:
                params = (row, row_num, ws, email2nickname, unknown_user_set, unknown_cell_list, dtable_uuid, repo_id,
                          image_param, cols_without_hidden, is_big_data_view)
                row_cells = handle_row(*params)
                ws.row_dimensions[row_num + 1].height = height_transfer(row_height)
            except Exception as e:
                if not row_error_log_exists:
                    dtable_io_logger.exception(e)
                    dtable_io_logger.error('Error row in exporting excel: {}'.format(e))
                    row_error_log_exists = True
                continue
            row_list.append(row_cells)

    if unknown_cell_list:
        try:
            add_nickname_to_cell(unknown_user_set, unknown_cell_list)
        except Exception as e:
            dtable_io_logger.exception('add nickname to cell error: {}'.format(e))
    for row in row_list:
        ws.append(row)


def handle_grouped_view_rows(view_rows, row_num_info, ws, email2nickname, unknown_user_set, unknown_cell_list, dtable_uuid,
                    repo_id, image_param, cols_without_hidden, column_name_to_column, summary_col_info, row_list, sub_level):
    head_name_to_head = {head.get('name'): head for head in cols_without_hidden}
    for row in view_rows:
        group_subgroups = row.get('subgroups')
        group_rows = row.get('rows')
        summaries = row.get('summaries', {})

        # write grouped row to ws
        row_cells = handle_grouped_row(row, ws, cols_without_hidden, column_name_to_column, sub_level, summary_col_info, head_name_to_head, summaries)
        row_num_info['row_num'] += 1
        row_list.append(row_cells)

        if group_rows is None and group_subgroups:
            handle_grouped_view_rows(group_subgroups, row_num_info, ws, email2nickname, unknown_user_set, unknown_cell_list,
                            dtable_uuid, repo_id, image_param, cols_without_hidden, column_name_to_column, summary_col_info, row_list, sub_level + 1)
        else:
            for group_row in group_rows:
                # write normal row to ws
                row_cells = handle_row(group_row, row_num_info.get('row_num'), ws, email2nickname, unknown_user_set, unknown_cell_list,
                                       dtable_uuid, repo_id, image_param, cols_without_hidden)
                row_list.append(row_cells)
                row_num_info['row_num'] += 1
