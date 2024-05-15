import base64
import json
import os
import shutil
import time
import uuid

import requests
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr, parseaddr
from urllib import parse
from datetime import datetime

from seaserv import seafile_api

from dtable_events.app.config import DTABLE_WEB_SERVICE_URL
from dtable_events.dtable_io.big_data import import_excel_to_db, update_excel_to_db, export_big_data_to_excel
from dtable_events.dtable_io.utils import setup_logger, \
    prepare_asset_file_folder, post_dtable_json, post_asset_files, \
    download_files_to_path, create_forms_from_src_dtable, copy_src_forms_to_json, \
    prepare_dtable_json_from_memory, update_page_design_static_image, \
    copy_src_auto_rules_to_json, create_auto_rules_from_src_dtable, sync_app_users_to_table, \
    copy_src_workflows_to_json, create_workflows_from_src_dtable, copy_src_external_app_to_json,\
    create_external_apps_from_src_dtable, zip_big_data_screen, post_big_data_screen_zip_file, \
    export_page_design_dir_to_path, update_page_design_content_to_path, upload_page_design, \
    download_page_design_file
from dtable_events.db import init_db_session_class
from dtable_events.dtable_io.excel import parse_excel_csv_to_json, import_excel_csv_by_dtable_server, \
    append_parsed_file_by_dtable_server, parse_append_excel_csv_upload_file_to_json, \
    import_excel_csv_add_table_by_dtable_server, update_parsed_file_by_dtable_server, \
    parse_update_excel_upload_excel_to_json, parse_update_csv_upload_csv_to_json, parse_and_import_excel_csv_to_dtable, \
    parse_and_import_excel_csv_to_table, parse_and_update_file_to_table, parse_and_append_excel_csv_to_table
from dtable_events.dtable_io.task_manager import task_manager
from dtable_events.page_design.utils import convert_page_to_pdf as _convert_page_to_pdf
from dtable_events.statistics.db import save_email_sending_records, batch_save_email_sending_records
from dtable_events.data_sync.data_sync_utils import run_sync_emails
from dtable_events.utils import get_inner_dtable_server_url, is_valid_email, uuid_str_to_36_chars
from dtable_events.utils.dtable_server_api import DTableServerAPI
from dtable_events.utils.exception import BaseSizeExceedsLimitError, ExcelFormatError
from dtable_events.dtable_io.utils import clear_tmp_dir, clear_tmp_file, clear_tmp_files_and_dirs

dtable_io_logger = setup_logger('dtable_events_io.log')
dtable_message_logger = setup_logger('dtable_events_message.log')
dtable_data_sync_logger = setup_logger('dtable_events_data_sync.log')
dtable_plugin_email_logger = setup_logger('dtable_events_plugin_email.log')


def get_dtable_export_content(username, repo_id, workspace_id, dtable_uuid, asset_dir_id, config):
    """
    1. prepare file content at /tmp/dtable-io/<dtable_id>/dtable_asset/...
    2. make zip file
    3. return zip file's content
    """
    dtable_io_logger.info('Start prepare /tmp/dtable-io/{}/zip_file.zip for export DTable.'.format(dtable_uuid))

    tmp_file_path = os.path.join('/tmp/dtable-io', dtable_uuid,
                                 'dtable_asset/')  # used to store asset files and json from file_server
    tmp_zip_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'zip_file') + '.zip'  # zip path of zipped xxx.dtable

    try:
        db_session = init_db_session_class(config)()
    except Exception as e:
        db_session = None
        dtable_io_logger.error('create db session failed. ERROR: {}'.format(e))
        raise Exception('create db session failed. ERROR: {}'.format(e))

    dtable_io_logger.info('Clear tmp dirs and files before prepare.')
    clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
    os.makedirs(tmp_file_path, exist_ok=True)
    # import here to avoid circular dependency

    # 1. create 'content.json' from 'xxx.dtable'
    dtable_io_logger.info('Create content.json file.')
    try:
        prepare_dtable_json_from_memory(workspace_id, dtable_uuid, username)
    except Exception as e:
        dtable_io_logger.error('prepare dtable json failed. ERROR: {}'.format(e))
        raise Exception('prepare dtable json failed. ERROR: {}'.format(e))

    # 2. get asset file folder, asset could be empty
    if asset_dir_id:
        dtable_io_logger.info('Create asset dir.')
        try:
            prepare_asset_file_folder(username, repo_id, dtable_uuid, asset_dir_id)
        except Exception as e:
            dtable_io_logger.warning('create asset folder failed. ERROR: {}'.format(e))

    # 3. copy forms
    try:
        copy_src_forms_to_json(dtable_uuid, tmp_file_path, db_session)
    except Exception as e:
        dtable_io_logger.error('copy forms failed. ERROR: {}'.format(e))
        raise Exception('copy forms failed. ERROR: {}'.format(e))
    finally:
        if db_session:
            db_session.close()

    # 4. copy automation rules
    try:
        copy_src_auto_rules_to_json(dtable_uuid, tmp_file_path, db_session)
    except Exception as e:
        dtable_io_logger.error('copy automation rules failed. ERROR: {}'.format(e))
        raise Exception('copy automation rules failed. ERROR: {}'.format(e))
    finally:
        if db_session:
            db_session.close()

    # 5. copy workflows
    try:
        copy_src_workflows_to_json(dtable_uuid, tmp_file_path, db_session)
    except Exception as e:
        dtable_io_logger.error('copy workflows failed. ERROR: {}'.format(e))
        raise Exception('copy workflows failed. ERROR: {}'.format(e))
    finally:
        if db_session:
            db_session.close()

    # 5. copy external app
    try:
        copy_src_external_app_to_json(dtable_uuid, tmp_file_path, db_session)
    except Exception as e:
        dtable_io_logger.error('copy external apps failed. ERROR: {}'.format(e))
        raise Exception('copy external apps failed. ERROR: {}'.format(e))
    finally:
        if db_session:
            db_session.close()

    """
    /tmp/dtable-io/<dtable_uuid>/dtable_asset/
                                    |- asset/
                                    |- content.json

    we zip /tmp/dtable-io/<dtable_uuid>/dtable_asset/ to /tmp/dtable-io/<dtable_id>/zip_file.zip and download it
    notice than make_archive will auto add .zip suffix to /tmp/dtable-io/<dtable_id>/zip_file
    """
    dtable_io_logger.info('Make zip file for download...')
    try:
        shutil.make_archive('/tmp/dtable-io/' + dtable_uuid + '/zip_file', "zip", root_dir=tmp_file_path)
    except Exception as e:
        dtable_io_logger.error('make zip failed. ERROR: {}'.format(e))
        raise Exception('make zip failed. ERROR: {}'.format(e))

    dtable_io_logger.info('Create /tmp/dtable-io/{}/zip_file.zip success!'.format(dtable_uuid))
    # we remove '/tmp/dtable-io/<dtable_uuid>' in dtable web api


def post_dtable_import_files(username, repo_id, workspace_id, dtable_uuid, dtable_file_name, in_storage,
                             can_use_automation_rules, can_use_workflows, can_use_external_apps, owner, org_id, config):
    """
    post files at /tmp/<dtable_uuid>/dtable_zip_extracted/ to file server
    unzip django uploaded tmp file is suppose to be done in dtable-web api.
    """
    dtable_io_logger.info('Start import DTable: {}.'.format(dtable_uuid))

    try:
        db_session = init_db_session_class(config)()
    except Exception as e:
        db_session = None
        dtable_io_logger.error('create db session failed. ERROR: {}'.format(e))

    dtable_io_logger.info('Prepare dtable json file and post it at file server.')
    try:
        dtable_content = post_dtable_json(username, repo_id, workspace_id, dtable_uuid, dtable_file_name, in_storage, db_session)
    except Exception as e:
        dtable_io_logger.error('post dtable json failed. ERROR: {}'.format(e))

    dtable_io_logger.info('Post asset files in tmp path to file server.')
    try:
        post_asset_files(repo_id, dtable_uuid, username)
    except Exception as e:
        dtable_io_logger.error('post asset files failed. ERROR: {}'.format(e))

    dtable_io_logger.info('create forms from src dtable.')
    try:
        create_forms_from_src_dtable(workspace_id, dtable_uuid, db_session)
    except Exception as e:
        dtable_io_logger.error('create forms failed. ERROR: {}'.format(e))
    finally:
        if db_session:
            db_session.close()

    old_new_workflow_token_dict = {}  # old new workflow token dict
    if can_use_workflows:
        dtable_io_logger.info('create workflows from src dtable.')
        try:
            create_workflows_from_src_dtable(username, workspace_id, repo_id, dtable_uuid, owner, org_id, old_new_workflow_token_dict, db_session)
        except Exception as e:
            dtable_io_logger.error('create workflows failed. ERROR: {}'.format(e))
        finally:
            if db_session:
                db_session.close()

    if can_use_automation_rules:
        dtable_io_logger.info('create auto rules from src dtable.')
        try:
            create_auto_rules_from_src_dtable(username, workspace_id, repo_id, owner, org_id, dtable_uuid, old_new_workflow_token_dict, db_session)
        except Exception as e:
            dtable_io_logger.error('create auto rules failed. ERROR: {}'.format(e))
        finally:
            if db_session:
                db_session.close()

    if can_use_external_apps:
        dtable_io_logger.info('create external apps from src dtable.')
        try:
            create_external_apps_from_src_dtable(username, dtable_uuid, db_session, org_id, workspace_id, repo_id)
        except Exception as e:
            dtable_io_logger.exception('create external apps failed. ERROR: {}'.format(e))
        finally:
            if db_session:
                db_session.close()

    try:
        if dtable_content:
            plugin_settings = dtable_content.get('plugin_settings', {})
            page_design_settings = plugin_settings.get('page-design', [])
            page_design_content_json_tmp_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'page-design')
            # handle different url in settings.py
            update_page_design_static_image(page_design_settings, repo_id, workspace_id, dtable_uuid, page_design_content_json_tmp_path, username)
    except Exception as e:
        dtable_io_logger.exception('update page design static image failed. ERROR: {}'.format(e))

    # remove extracted tmp file
    dtable_io_logger.info('Remove extracted tmp file.')
    try:
        shutil.rmtree(os.path.join('/tmp/dtable-io', dtable_uuid))
    except Exception as e:
        dtable_io_logger.error('rm extracted tmp file failed. ERROR: {}'.format(e))

    dtable_io_logger.info('Import DTable: {} success!'.format(dtable_uuid))

def get_dtable_export_asset_files(username, repo_id, dtable_uuid, files, task_id, config, files_map=None):
    """
    export asset files from dtable
    """
    handled_files = []
    for file in files:
        files_map_value = files_map.get(file) if files_map else None
        if files_map_value:
            files_map.pop(file)
        file = file.strip().strip('/')
        file = file[:file.find('?')] if '?' in file else file
        handled_files.append(file)
        if files_map_value:
            files_map[file] = files_map_value
    files = handled_files
    tmp_file_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'asset-files',
                                 str(task_id))           # used to store files
    tmp_zip_path  = os.path.join('/tmp/dtable-io', dtable_uuid, 'asset-files',
                                 str(task_id)) + '.zip'  # zip those files

    clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
    os.makedirs(tmp_file_path, exist_ok=True)

    db_session = init_db_session_class(config)()
    try:
        # 1. download files to tmp_file_path
        download_files_to_path(username, repo_id, dtable_uuid, files, tmp_file_path, db_session, files_map)
        # 2. zip those files to tmp_zip_path
        shutil.make_archive(tmp_zip_path.split('.')[0], 'zip', root_dir=tmp_file_path)
    except Exception as e:
        dtable_io_logger.error('export asset files from dtable failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('export files from dtable: %s success!', dtable_uuid)
    finally:
        db_session.close()

def get_dtable_export_big_data_screen(username, repo_id, dtable_uuid, page_id, task_id):
    """
    parse json file in big data screen, and zip it for download
    """
    tmp_file_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'big-data-screen', str(task_id))
    tmp_zip_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'big-data-screen', str(task_id) + '.zip')
    clear_tmp_files_and_dirs(tmp_file_path, tmp_zip_path)
    os.makedirs(tmp_file_path.rstrip('/') + '/images', exist_ok=True)

    try:
        zip_big_data_screen(username, repo_id, dtable_uuid, page_id, tmp_file_path)
        shutil.make_archive(tmp_zip_path.split('.')[0], 'zip', root_dir=tmp_file_path)
    except Exception as e:
        dtable_io_logger.error('export big data screen from dtable failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('export big data screen from dtable: %s success!', dtable_uuid)

def import_big_data_screen(username, repo_id, dtable_uuid, page_id):
    """
    parse the zip in tmp folders and upload it
    """
    tmp_extracted_path = os.path.join('/tmp/dtable-io', dtable_uuid, 'big_data_screen_zip_extracted/')
    try:
        post_big_data_screen_zip_file(username, repo_id, dtable_uuid, page_id, tmp_extracted_path)
    except Exception as e:
        dtable_io_logger.error('import big data screen from dtable failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('import big data screen to dtable: %s success!', dtable_uuid)
    try:
        shutil.rmtree(tmp_extracted_path)
    except Exception as e:
        dtable_io_logger.error('rm extracted tmp file failed. ERROR: {}'.format(e))

def parse_excel_csv(username, repo_id, file_name, file_type, parse_type, dtable_uuid, config):
    """
    parse excel or csv to json file, then upload json file to file server
    """
    dtable_io_logger.info('Start parse excel or csv: %s.%s.' % (file_name, file_type))
    try:
        parse_excel_csv_to_json(username, repo_id, file_name, file_type, parse_type, dtable_uuid)
    except ExcelFormatError as e:
        raise Exception(e)
    except Exception as e:
        dtable_io_logger.exception('parse excel or csv failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('parse excel %s.xlsx success!' % file_name)

def import_excel_csv(username, repo_id, dtable_uuid, dtable_name, included_tables, lang, config):
    """
    upload excel or csv json file to dtable-server
    """
    dtable_io_logger.info('Start import excel or csv: {}.'.format(dtable_uuid))
    try:
        import_excel_csv_by_dtable_server(username, repo_id, dtable_uuid, dtable_name, included_tables, lang)
    except BaseSizeExceedsLimitError:
        raise Exception('Base size exceeds limit')
    except Exception as e:
        dtable_io_logger.error('import excel or csv failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('import excel or csv %s success!' % dtable_name)

def import_excel_csv_add_table(username, dtable_uuid, dtable_name, included_tables, lang, config):
    """
    add table, upload excel or csv json file to dtable-server
    """
    dtable_io_logger.info('Start import excel or csv add table: {}.'.format(dtable_uuid))
    try:
        import_excel_csv_add_table_by_dtable_server(username, dtable_uuid, dtable_name, included_tables, lang)
    except BaseSizeExceedsLimitError:
        raise Exception('Base size exceeds limit')
    except Exception as e:
        dtable_io_logger.error('import excel or csv add table failed. dtable_uuid: %s, dtable_name: %s ERROR: %s' % (dtable_uuid, dtable_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('import excel or csv %s add table success!' % dtable_name)


def append_excel_csv_append_parsed_file(username, dtable_uuid, file_name, table_name):
    """
    upload excel or csv json file to dtable-server
    """
    dtable_io_logger.info('Start import excel or csv: {}.'.format(dtable_uuid))
    try:
        append_parsed_file_by_dtable_server(username, dtable_uuid, file_name, table_name)
    except Exception as e:
        dtable_io_logger.exception('append excel or csv failed. dtable_uuid: %s, table_name: %s ERROR:  %s' % (dtable_uuid, table_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('append excel or csv %s success!' % file_name)

def append_excel_csv_upload_file(username, file_name, dtable_uuid, table_name, file_type):
    """
    parse excel or csv to json file, then upload json file to file server
    """
    dtable_io_logger.info('Start parse append excel or csv: %s.%s' % (file_name, file_type))
    try:
        parse_append_excel_csv_upload_file_to_json(file_name, username, dtable_uuid, table_name, file_type)
    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('parse append excel or csv failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('parse append excel or csv %s.%s success!' % (file_name, file_type))

def update_excel_csv_update_parsed_file(username, dtable_uuid, file_name, table_name, selected_columns):
    """
    upload excel/csv json file to dtable-server
    """
    dtable_io_logger.info('Start import file: {}.'.format(dtable_uuid))
    try:
        update_parsed_file_by_dtable_server(username, dtable_uuid, file_name, table_name, selected_columns)
    except Exception as e:
        dtable_io_logger.exception('update excel,csv failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Update excel or csv error')
    else:
        dtable_io_logger.info('update excel,csv %s success!' % file_name)

def update_excel_upload_excel(username, file_name, dtable_uuid, table_name):
    """
    parse excel to json file, then upload json file to file server
    """
    dtable_io_logger.info('Start parse update excel: %s.xlsx.' % file_name)
    try:
        parse_update_excel_upload_excel_to_json(file_name, username, dtable_uuid, table_name)
    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('parse update excel failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Update excel or csv error')
    else:
        dtable_io_logger.info('parse update excel %s.xlsx success!' % file_name)

def update_csv_upload_csv(username, file_name, dtable_uuid, table_name):
    """
    parse csv to json file, then upload json file to file server
    """
    dtable_io_logger.info('Start parse update csv: %s.csv.' % file_name)
    try:
        parse_update_csv_upload_csv_to_json(file_name, username, dtable_uuid, table_name)
    except Exception as e:
        dtable_io_logger.exception('parse update csv failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Update excel or csv error')
    else:
        dtable_io_logger.info('parse update csv %s.csv success!' % file_name)


def import_excel_csv_to_dtable(username, repo_id, dtable_name, dtable_uuid, file_type, lang):
    """
    parse excel csv to json, then import excel csv to dtable
    """
    dtable_io_logger.info('Start import excel or csv: %s.%s to dtable.' % (dtable_name, file_type))
    try:
        parse_and_import_excel_csv_to_dtable(repo_id, dtable_name, dtable_uuid, username, file_type, lang)
    except BaseSizeExceedsLimitError:
        raise Exception('Base size exceeds limit')
    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('import excel or csv to dtable failed. dtable_uuid: %s, dtable_name: %s ERROR: %s' % (dtable_uuid, dtable_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('import excel or csv %s.%s to dtable success!' % (dtable_name, file_type))


def import_excel_csv_to_table(username, file_name, dtable_uuid, file_type, lang):
    """
    parse excel or csv to json, then import excel or csv to table
    """
    dtable_io_logger.info('Start import excel or csv: %s.%s to table.' % (file_name, file_type))
    try:
        parse_and_import_excel_csv_to_table(file_name, dtable_uuid, username, file_type, lang)
    except BaseSizeExceedsLimitError:
        raise Exception('Base size exceeds limit')
    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('import excel or csv to table failed.  dtable_uuid: %s, file_name: %s ERROR: %s' % (dtable_uuid, file_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('import excel or csv %s.%s to table success!' % (file_name, file_type))


def update_table_via_excel_csv(username, file_name, dtable_uuid, table_name, selected_columns, file_type):
    """
    update excel/csv file to table
    """
    dtable_io_logger.info('Start update file: %s.%s to table.' % (file_name, file_type))
    try:
        parse_and_update_file_to_table(file_name, username, dtable_uuid, table_name, selected_columns, file_type)
    except Exception as e:
        dtable_io_logger.exception('update file update to table failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Update excel or csv error')
    else:
        dtable_io_logger.info('update file %s.%s update to table success!' % (file_name, file_type))


def append_excel_csv_to_table(username, file_name, dtable_uuid, table_name, file_type):
    """
    parse excel or csv to json, then append excel or csv to table
    """
    dtable_io_logger.info('Start append excel or csv: %s.%s to table.' % (file_name, file_type))
    try:
        parse_and_append_excel_csv_to_table(username, file_name, dtable_uuid, table_name, file_type)
    except ExcelFormatError:
        raise Exception('Excel format error')
    except Exception as e:
        dtable_io_logger.exception('append excel or csv to table failed. dtable_uuid: %s, table_name: %s ERROR: %s' % (dtable_uuid, table_name, e))
        raise Exception('Import excel or csv error')
    else:
        dtable_io_logger.info('append excel or csv %s.%s to table success!' % (file_name, file_type))


def _get_upload_link_to_seafile(seafile_server_url, access_token, parent_dir="/"):
    upload_link_api_url = "%s%s" % (seafile_server_url.rstrip('/'),  '/api/v2.1/via-repo-token/upload-link/')
    headers = {
        'authorization': 'Token ' + access_token
    }
    params = {
        'path': parent_dir
    }
    response = requests.get(upload_link_api_url, headers=headers, params=params)
    return response.json()

def _upload_to_seafile(seafile_server_url, access_token, files, parent_dir="/", relative_path="", replace=None):
    upload_url = _get_upload_link_to_seafile(seafile_server_url, access_token, parent_dir)
    files_tuple_list = [('file', open(file, 'rb')) for file in files]
    files = files_tuple_list + [('parent_dir', parent_dir), ('relative_path', relative_path), ('replace', replace)]
    response = requests.post(upload_url, files=files)
    return response

def get_dtable_transfer_asset_files(username, repo_id, dtable_uuid, files, task_id, files_map, parent_dir, relative_path, replace, repo_api_token, seafile_server_url, config):
    tmp_file_path = os.path.join('/tmp/dtable-io/', dtable_uuid, 'transfer-files', str(task_id))
    os.makedirs(tmp_file_path, exist_ok=True)

    db_session = init_db_session_class(config)()
    try:
        # download files to local
        local_file_list = download_files_to_path(username, repo_id, dtable_uuid, files, tmp_file_path, db_session, files_map)
    except Exception as e:
        dtable_io_logger.error('export asset files from dtable failed. ERROR: {}'.format(e))
        if os.path.exists(tmp_file_path):
            shutil.rmtree(tmp_file_path)
        return
    finally:
        db_session.close()

    # upload files from local to seafile
    try:
        _upload_to_seafile(seafile_server_url, repo_api_token, local_file_list, parent_dir, relative_path, replace)
    except Exception as e:
        dtable_io_logger.error('transfer asset files from dtable failed. ERROR: {}'.format(e))

    # delete local files
    if os.path.exists(tmp_file_path):
        shutil.rmtree(tmp_file_path)

def send_wechat_msg(webhook_url, msg, msg_type="text"):
    if msg_type == "markdown":
        msg_format = {"msgtype": "markdown","markdown":{"content":msg}}
    else:
        msg_format = {"msgtype": "text", "text": {"content": msg}}
    result = {}
    try:
        requests.post(webhook_url, json=msg_format, headers={"Content-Type": "application/json"})
    except Exception as e:
        dtable_message_logger.error('Wechat sending failed. ERROR: {}'.format(e))
        result['err_msg'] = 'Webhook URL invalid'
    else:
        dtable_message_logger.info('Wechat sending success!')
    return result

def send_dingtalk_msg(webhook_url, msg, msg_type="text", msg_title=None):
    result = {}
    if msg_type == "markdown":
        if not msg_title:
            result['err_msg'] = 'msg_title invalid'
            dtable_message_logger.error('Dingtalk sending failed. ERROR: msg_title invalid')
            return result
        msg_format = {"msgtype": "markdown", "markdown": {"text": msg, "title": msg_title}}
    else:
        msg_format = {"msgtype": "text", "text": {"content": msg}}

    try:
        requests.post(webhook_url, json=msg_format, headers={"Content-Type": "application/json"})
    except Exception as e:
        dtable_message_logger.error('Dingtalk sending failed. ERROR: {}'.format(e))
        result['err_msg'] = 'Dingtalk URL invalid'
    else:
        dtable_message_logger.info('Dingtalk sending success!')
    return result

def send_notification_msg(emails, user_col_key, msg, dtable_uuid, username, table_id=None, row_id=None):
    result = {}
    try:
        dtable_server_url = get_inner_dtable_server_url()
        dtable_server_api = DTableServerAPI(username, dtable_uuid, dtable_server_url)
        metadata = dtable_server_api.get_metadata()
        table = None
        for tmp_table in metadata['tables']:
            if tmp_table['_id'] == table_id:
                table = tmp_table
                break
        if not table:
            return
        
        target_row = dtable_server_api.get_row(table['name'], row_id)
        
        sending_list = emails
        if user_col_key:
            column = None
            for tmp_col in table['columns']:
                if tmp_col['key'] == user_col_key:
                    column = tmp_col
                    break

            user_col_info = column and target_row.get(column['name']) or None
            if user_col_info:
                if isinstance(user_col_info, list):
                    for user in user_col_info:
                        if user in sending_list:
                            continue
                        sending_list.append(user)
                else:
                    if user_col_info not in sending_list:
                        sending_list.append(user_col_info)

        detail = {
            'table_id': table_id or '',
            'msg': msg,
            'row_id_list': row_id and [row_id, ] or [],
        }
        user_msg_list = []
        for user in sending_list:
            if not is_valid_email(user):
                continue
            user_msg_list.append({
                'to_user': user,
                'msg_type': 'notification_rules',
                'detail': detail,
                })

        dtable_server_api.batch_send_notification(user_msg_list)
    except Exception as e:
        dtable_message_logger.error('Notification sending failed. ERROR: {}'.format(e))
        result['err_msg'] = 'Notification send failed'
    else:
        dtable_message_logger.info('Notification sending success!')
    return result


def send_email_msg(auth_info, send_info, username, config=None, db_session=None):
    # auth info
    email_host = auth_info.get('email_host')
    email_port = int(auth_info.get('email_port', 0))
    host_user = auth_info.get('host_user')
    password = auth_info.get('password')
    sender_name = auth_info.get('sender_name', '')

    # send info
    msg = send_info.get('message', '')
    html_msg = send_info.get('html_message', '')
    send_to = send_info.get('send_to', [])
    subject = send_info.get('subject', '')
    source = send_info.get('source', '')
    copy_to = send_info.get('copy_to', [])
    reply_to = send_info.get('reply_to', '')
    file_download_urls = send_info.get('file_download_urls', None)
    message_id = send_info.get('message_id', '')
    in_reply_to = send_info.get('in_reply_to', '')
    image_cid_url_map = send_info.get('image_cid_url_map', {})

    send_to = [formataddr(parseaddr(to)) for to in send_to]
    copy_to = [formataddr(parseaddr(to)) for to in copy_to]
    if source:
        source = formataddr(parseaddr(source))

    result = {}
    if not msg and not html_msg:
        result['err_msg'] = 'Email message invalid'
        return result

    msg_obj = MIMEMultipart()
    msg_obj['Subject'] = subject
    msg_obj['From'] = source or formataddr((sender_name, host_user))
    msg_obj['To'] = ",".join(send_to)
    msg_obj['Cc'] = ",".join(copy_to)
    msg_obj['Reply-to'] = reply_to

    if message_id:
        msg_obj['Message-ID'] = message_id

    if in_reply_to:
        msg_obj['In-Reply-To'] = in_reply_to

    if msg:
        plain_content_body = MIMEText(msg)
        msg_obj.attach(plain_content_body)

    if html_msg:
        html_content_body = MIMEText(html_msg, 'html')
        msg_obj.attach(html_content_body)

    if html_msg and image_cid_url_map:
        for cid, image_url in image_cid_url_map.items():
            response = requests.get(image_url)
            from email.mime.image import MIMEImage
            msg_image = MIMEImage(response.content)
            msg_image.add_header('Content-ID', '<%s>' % cid)
            msg_obj.attach(msg_image)

    if file_download_urls:
        for file_name, file_url in file_download_urls.items():
            response = requests.get(file_url)
            attach_file = MIMEText(response.content, 'base64', 'utf-8')
            attach_file["Content-Type"] = 'application/octet-stream'
            attach_file["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'' + parse.quote(file_name)
            msg_obj.attach(attach_file)

    try:
        smtp = smtplib.SMTP(email_host, int(email_port), timeout=30)
    except Exception as e:
        dtable_message_logger.warning(
            'Email server configured failed. host: %s, port: %s, error: %s' % (email_host, email_port, e))
        result['err_msg'] = 'Email server host or port invalid'
        return result
    success = False

    try:
        smtp.starttls()
        smtp.login(host_user, password)
        recevers = copy_to and send_to + copy_to or send_to
        smtp.sendmail(host_user, recevers, msg_obj.as_string())
        success = True
    except Exception as e:
        dtable_message_logger.warning(
            'Email sending failed. email: %s, error: %s' % (host_user, e))
        result['err_msg'] = 'Email server username or password invalid'
    else:
        dtable_message_logger.info('Email sending success!')
    finally:
        smtp.quit()

    session = db_session or init_db_session_class(config)()
    try:
        save_email_sending_records(session, username, email_host, success)
    except Exception as e:
        dtable_message_logger.error(
            'Email sending log record error: %s' % e)
    finally:
        session.close()
    return result


def batch_send_email_msg(auth_info, send_info_list, username, config=None, db_session=None):
    """
    for personal user of email, this function only support sending 10 emails per time
    """
    # auth info
    email_host = auth_info.get('email_host')
    email_port = int(auth_info.get('email_port'))
    host_user = auth_info.get('host_user')
    password = auth_info.get('password')
    sender_name = auth_info.get('sender_name', '')

    try:
        smtp = smtplib.SMTP(email_host, int(email_port), timeout=30)
    except Exception as e:
        dtable_message_logger.warning(
            'Email server configured failed. host: %s, port: %s, error: %s' % (email_host, email_port, e))
        return

    try:
        smtp.starttls()
        smtp.login(host_user, password)
    except Exception as e:
        dtable_message_logger.warning(
            'Login smtp failed, host user: %s, error: %s' % (host_user, e))
        return

    send_state_list = []
    for send_info in send_info_list:
        success = False
        msg = send_info.get('message', '')
        html_msg = send_info.get('html_message', '')
        send_to = send_info.get('send_to', [])
        subject = send_info.get('subject', '')
        source = send_info.get('source', '')
        copy_to = send_info.get('copy_to', [])
        reply_to = send_info.get('reply_to', '')
        file_download_urls = send_info.get('file_download_urls', None)
        image_cid_url_map = send_info.get('image_cid_url_map', {})

        if not msg and not html_msg:
            dtable_message_logger.warning('Email message invalid')
            continue

        send_to = [formataddr(parseaddr(to)) for to in send_to]
        copy_to = [formataddr(parseaddr(to)) for to in copy_to]

        msg_obj = MIMEMultipart()
        msg_obj['Subject'] = subject
        msg_obj['From'] = source or formataddr((sender_name, host_user))
        msg_obj['To'] = ",".join(send_to)
        msg_obj['Cc'] = copy_to and ",".join(copy_to) or ""
        msg_obj['Reply-to'] = reply_to

        if msg:
            plain_content_body = MIMEText(msg)
            msg_obj.attach(plain_content_body)

        if html_msg:
            html_content_body = MIMEText(html_msg, 'html')
            msg_obj.attach(html_content_body)

        if html_msg and image_cid_url_map:
            for cid, image_url in image_cid_url_map.items():
                response = requests.get(image_url)
                from email.mime.image import MIMEImage
                msg_image = MIMEImage(response.content)
                msg_image.add_header('Content-ID', '<%s>' % cid)
                msg_obj.attach(msg_image)

        if file_download_urls:
            for file_name, file_url in file_download_urls.items():
                response = requests.get(file_url)
                attach_file = MIMEText(response.content, 'base64', 'utf-8')
                attach_file["Content-Type"] = 'application/octet-stream'
                attach_file["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'' + parse.quote(file_name)
                msg_obj.attach(attach_file)

        try:
            recevers = copy_to and send_to + copy_to or send_to
            smtp.sendmail(host_user, recevers, msg_obj.as_string())
            success = True
        except Exception as e:
            dtable_message_logger.warning('Email sending failed. email: %s, error: %s' % (host_user, e))
        else:
            dtable_message_logger.info('Email sending success!')
        send_state_list.append(success)
        time.sleep(0.5)

    smtp.quit()

    session = db_session or init_db_session_class(config)()
    try:
        batch_save_email_sending_records(session, username, email_host, send_state_list)
    except Exception as e:
        dtable_message_logger.error('Batch save email sending log error: %s' % e)
    finally:
        session.close()


def convert_page_to_pdf(dtable_uuid, page_id, row_id):
    dtable_server_url = get_inner_dtable_server_url()
    access_token = DTableServerAPI('dtable-events', dtable_uuid, dtable_server_url).internal_access_token
    target_dir = '/tmp/dtable-io/convert-page-to-pdf'
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)
    target_path = os.path.join(target_dir, '%s_%s_%s.pdf' % (dtable_uuid, page_id, row_id))

    _convert_page_to_pdf(dtable_uuid, page_id, row_id, access_token, target_path)


def convert_view_to_execl(dtable_uuid, table_id, view_id, username, id_in_org, user_department_ids_map, permission, name, repo_id, is_support_image=False):
    from dtable_events.dtable_io.utils import get_metadata_from_dtable_server, get_view_rows_from_dtable_server
    from dtable_events.dtable_io.excel import write_xls_with_type, TEMP_EXPORT_VIEW_DIR, IMAGE_TMP_DIR
    from dtable_events.dtable_io.utils import get_related_nicknames_from_dtable, escape_sheet_name
    import openpyxl

    target_dir = TEMP_EXPORT_VIEW_DIR + dtable_uuid
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)

    try:
        nicknames = get_related_nicknames_from_dtable(dtable_uuid, username, permission)
    except Exception as e:
        dtable_io_logger.error('get nicknames. ERROR: {}'.format(e))
        return
    email2nickname = {nickname['email']: nickname['name'] for nickname in nicknames}

    try:
        metadata = get_metadata_from_dtable_server(dtable_uuid, username)
    except Exception as e:
        dtable_io_logger.error('get metadata. ERROR: {}'.format(e))
        return

    target_table = {}
    target_view = {}
    for table in metadata.get('tables', []):
        if table.get('_id', '') == table_id:
            target_table = table
            break

    if not target_table:
        dtable_io_logger.warning('Table %s not found.' % table_id)
        return

    for view in target_table.get('views', []):
        if view.get('_id', '') == view_id:
            target_view = view
            break
    if not target_view:
        dtable_io_logger.warning('View %s not found.' % view_id)
        return

    table_name = target_table.get('name', '')
    view_name = target_view.get('name', '')
    row_height = target_view.get('row_height', 'default')
    header_height = 'default'
    header_settings = target_table.get('header_settings')
    if header_settings:
        header_height = header_settings.get('header_height', 'default')

    cols = target_table.get('columns', [])
    hidden_cols_key = target_view.get('hidden_columns', [])
    summary_configs = target_table.get('summary_configs', {})
    cols_without_hidden = []
    summary_col_info = {}
    for col in cols:
        if col.get('key', '') not in hidden_cols_key:
            cols_without_hidden.append(col)
        if summary_configs.get(col.get('key')):
            summary_col_info.update({col.get('name'): summary_configs.get(col.get('key'))})

    images_target_dir = os.path.join(IMAGE_TMP_DIR, dtable_uuid, str(uuid.uuid4()))
    image_param = {'num': 0, 'is_support': is_support_image, 'images_target_dir': images_target_dir}

    sheet_name = table_name + ('_' + view_name if view_name else '')
    sheet_name = escape_sheet_name(sheet_name)
    excel_name = name + '_' + table_name + ('_' + view_name if view_name else '') + '.xlsx'

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)

    res_json = get_view_rows_from_dtable_server(dtable_uuid, table_id, view_id, username, id_in_org, user_department_ids_map, permission, table_name, view_name)
    dtable_rows = res_json.get('rows', [])

    column_name_to_column = {col.get('name'): col for col in cols}
    is_group_view = bool(target_view.get('groupbys'))

    params = (dtable_rows, email2nickname, ws, 0, dtable_uuid, repo_id, image_param, cols_without_hidden, column_name_to_column, is_group_view, summary_col_info, row_height, header_height)

    try:
        write_xls_with_type(*params)
    except Exception as e:
        dtable_io_logger.exception(e)
        dtable_io_logger.error('head_list = {}\n{}'.format(cols_without_hidden, e))
        return
    target_path = os.path.join(target_dir, excel_name)
    wb.save(target_path)
    # remove tmp images
    try:
        shutil.rmtree(images_target_dir)
    except:
        pass


def convert_table_to_execl(dtable_uuid, table_id, username, permission, name, repo_id, is_support_image=False):
    from dtable_events.dtable_io.utils import get_metadata_from_dtable_server, get_rows_from_dtable_server
    from dtable_events.dtable_io.excel import write_xls_with_type, IMAGE_TMP_DIR
    from dtable_events.dtable_io.utils import get_related_nicknames_from_dtable, escape_sheet_name
    import openpyxl

    target_dir = '/tmp/dtable-io/export-table-to-excel/' + dtable_uuid
    if not os.path.isdir(target_dir):
        os.makedirs(target_dir)

    try:
        nicknames = get_related_nicknames_from_dtable(dtable_uuid, username, permission)
    except Exception as e:
        dtable_io_logger.error('get nicknames. ERROR: {}'.format(e))
        return
    email2nickname = {nickname['email']: nickname['name'] for nickname in nicknames}

    try:
        metadata = get_metadata_from_dtable_server(dtable_uuid, username)
    except Exception as e:
        dtable_io_logger.error('get metadata. ERROR: {}'.format(e))
        return

    target_table = {}
    for table in metadata.get('tables', []):
        if table.get('_id', '') == table_id:
            target_table = table
            break

    if not target_table:
        dtable_io_logger.warning('Table %s not found.' % table_id)
        return

    table_name = target_table.get('name', '')
    cols = target_table.get('columns', [])
    header_height = 'default'
    header_settings = target_table.get('header_settings')
    if header_settings:
        header_height = header_settings.get('header_height', 'default')

    result_rows = get_rows_from_dtable_server(username, dtable_uuid, table_name)
    column_name_to_column = {col.get('name'): col for col in cols}

    images_target_dir = os.path.join(IMAGE_TMP_DIR, dtable_uuid, str(uuid.uuid4()))
    image_param = {'num': 0, 'is_support': is_support_image, 'images_target_dir': images_target_dir}

    sheet_name = escape_sheet_name(table_name)
    excel_name = name + '_' + table_name + '.xlsx'
    target_path = os.path.join(target_dir, excel_name)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)
    try:
        write_xls_with_type(result_rows, email2nickname, ws, 0, dtable_uuid, repo_id, image_param, cols, column_name_to_column, header_height=header_height)
    except Exception as e:
        dtable_io_logger.error('head_list = {}\n{}'.format(cols, e))
        return
    wb.save(target_path)
    # remove tmp images
    try:
        shutil.rmtree(images_target_dir)
    except:
        pass

def app_user_sync(dtable_uuid, app_name, app_id, table_name, table_id, username, config):
    dtable_io_logger.info('Start sync app %s users: to table %s.' % (app_name, table_name))
    db_session = init_db_session_class(config)()
    try:
        sync_app_users_to_table(dtable_uuid, app_id, table_name, table_id, username, db_session)
    except Exception as e:

        dtable_io_logger.exception('app user sync ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('app %s user sync success!' % app_name)
    finally:
        if db_session:
            db_session.close()


def email_sync(context, config):
    dtable_data_sync_logger.info('Start sync email to dtable %s, email table %s.' % (context.get('dtable_uuid'), context.get('detail',{}).get('email_table_id')))
    db_session = init_db_session_class(config)()
    context['db_session'] = db_session

    try:
        run_sync_emails(context)
    except Exception as e:
        dtable_data_sync_logger.exception('sync email ERROR: {}'.format(e))
    else:
        dtable_data_sync_logger.info('sync email success, sync_id: %s' % context.get('data_sync_id'))
    finally:
        if db_session:
            db_session.close()


def plugin_email_send_email(context, config=None):
    dtable_plugin_email_logger.info('Start send email by plugin %s, email table %s.' % (context.get('dtable_uuid'), context.get('table_info', {}).get('email_table_name')))

    dtable_uuid = context.get('dtable_uuid')
    username = context.get('username')
    repo_id = context.get('repo_id')
    workspace_id = context.get('workspace_id')

    table_info = context.get('table_info')
    email_info = context.get('email_info')
    auth_info = context.get('auth_info')

    thread_row_id = table_info.get('thread_row_id')
    email_row_id = table_info.get('email_row_id')
    email_table_name = table_info.get('email_table_name')
    thread_table_name = table_info.get('thread_table_name')

    # send email
    result = send_email_msg(auth_info, email_info, username, config)

    if result.get('err_msg'):
        dtable_plugin_email_logger.error('plugin email send failed, email account: %s, username: %s', auth_info.get('host_user'), username)
        return

    send_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    api_url = get_inner_dtable_server_url()
    dtable_server_api = DTableServerAPI(username, dtable_uuid, api_url, server_url=DTABLE_WEB_SERVICE_URL,
                                        repo_id=repo_id, workspace_id=workspace_id)

    replied_email_row = dtable_server_api.get_row(email_table_name, email_row_id)

    thread_id = replied_email_row.get('Thread ID')

    html_message = email_info.get('html_message')
    if html_message:
        html_message = '```' + html_message + '```'

    email = {
        'cc': email_info.get('copy_to'),
        'From': email_info.get('from'),
        'Message ID': email_info.get('message_id'),
        'Reply to Message ID': email_info.get('in_reply_to'),
        'To': email_info.get('send_to'),
        'Subject': email_info.get('subject'),
        'Content': email_info.get('text_message'),
        'HTML Content': html_message,
        'Date': send_time,
        'Thread ID': thread_id,
    }

    metadata = dtable_server_api.get_metadata()

    tables = metadata.get('tables', [])
    email_table_id = ''
    link_table_id = ''
    for table in tables:
        if table.get('name') == email_table_name:
            email_table_id = table.get('_id')
        if table.get('name') == thread_table_name:
            link_table_id = table.get('_id')
        if email_table_id and link_table_id:
            break
    if not email_table_id or not link_table_id:
        dtable_plugin_email_logger.error('email table: %s or link table: %s not found', email_table_name, thread_table_name)
        return

    email_link_id = dtable_server_api.get_column_link_id(email_table_name, 'Threads')

    email_row = dtable_server_api.append_row(email_table_name, email)
    email_row_id = email_row.get('_id')
    other_rows_ids = [thread_row_id]

    dtable_server_api.update_link(email_link_id, email_table_id, link_table_id, email_row_id, other_rows_ids)

    dtable_server_api.update_row(thread_table_name, thread_row_id, {'Last Updated': send_time})

def import_big_excel(username, dtable_uuid, table_name, file_path, task_id, tasks_status_map):
    """
    upload excel json file to dtable-db
    """

    dtable_io_logger.info('Start import big excel: {}.'.format(dtable_uuid))
    try:
        import_excel_to_db(username, dtable_uuid, table_name, file_path, task_id, tasks_status_map)
    except Exception as e:
        dtable_io_logger.error('import big excel failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('import big excel %s.xlsx success!' % table_name)


def update_big_excel(username, dtable_uuid, table_name, file_path, ref_columns, is_insert_new_data, task_id, tasks_status_map):
    """
    upload excel json file to dtable-db
    """

    dtable_io_logger.info('Start update big excel: {}.'.format(dtable_uuid))
    try:
        update_excel_to_db(username, dtable_uuid, table_name, file_path, ref_columns, is_insert_new_data, task_id, tasks_status_map)
    except Exception as e:
        dtable_io_logger.error('update big excel failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('update big excel %s.xlsx success!' % table_name)


def convert_big_data_view_to_execl(dtable_uuid, table_id, view_id, username, name, task_id, tasks_status_map, repo_id, is_support_image):
    dtable_io_logger.info('Start export big data view to excel: {}.'.format(dtable_uuid))
    try:
        export_big_data_to_excel(dtable_uuid, table_id, view_id, username, name, task_id, tasks_status_map, repo_id, is_support_image)
    except Exception as e:
        dtable_io_logger.error('export big data view failed. ERROR: {}'.format(e))
    else:
        dtable_io_logger.info('export big data table_id: %s, view_id: %s success!', table_id, view_id)


def export_page_design(repo_id, dtable_uuid, page_id, username):
    # prepare empty dir
    tmp_zip_dir = os.path.join('/tmp/dtable-io', 'page-design')
    os.makedirs(tmp_zip_dir, exist_ok=True)
    tmp_zip_path = os.path.join(tmp_zip_dir, f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.zip')

    # download and save to path
    export_page_design_dir_to_path(repo_id, dtable_uuid, page_id, tmp_zip_path, username)


def import_page_design(repo_id, workspace_id, dtable_uuid, page_id, is_dir, username):
    # check file exists
    need_check_static = False
    try:
        if is_dir:
            tmp_page_path = os.path.join('/tmp/dtable-io', 'page-design', f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}')
            download_page_design_file(repo_id, dtable_uuid, page_id, is_dir, username)
            items = os.listdir(tmp_page_path)
            if 'static_image' in items:
                need_check_static = True
        else:
            download_page_design_file(repo_id, dtable_uuid, page_id, is_dir, username)
            tmp_page_path = os.path.join('/tmp/dtable-io', 'page-design', f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.json')
    except Exception as e:
        if is_dir:
            tmp_page_path = os.path.join('/tmp/dtable-io', 'page-design', f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}')
            clear_tmp_dir(tmp_page_path)
            clear_tmp_file(tmp_page_path + '.zip')
        else:
            tmp_page_path = os.path.join('/tmp/dtable-io', 'page-design', f'{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.json')
            clear_tmp_file(tmp_page_path)
        raise e
    finally:
        if is_dir:
            try:
                seafile_tmp_file = f'/asset/{uuid_str_to_36_chars(dtable_uuid)}/page-design/{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.zip'
                if seafile_api.get_file_id_by_path(repo_id, seafile_tmp_file):
                    seafile_api.del_file(repo_id, os.path.dirname(seafile_tmp_file), os.path.basename(seafile_tmp_file), username)
            except Exception as e:
                dtable_io_logger.exception('delete repo: %s temp zip file: %s error: %s', repo_id, tmp_page_path, e)
        else:
            try:
                seafile_tmp_file = f'/asset/{uuid_str_to_36_chars(dtable_uuid)}/page-design/{uuid_str_to_36_chars(dtable_uuid)}-{page_id}.json'
                if seafile_api.get_file_id_by_path(repo_id, seafile_tmp_file):
                    seafile_api.del_file(repo_id, os.path.dirname(seafile_tmp_file), os.path.basename(seafile_tmp_file), username)
            except Exception as e:
                dtable_io_logger.exception('delete repo: %s temp zip file: %s error: %s', repo_id, tmp_page_path, e)

    if not os.path.exists(tmp_page_path):
        return

    try:
        if is_dir:
            # update content and save to file
            tmp_content_file = os.path.join(tmp_page_path, f'{page_id}.json')
            update_page_design_content_to_path(workspace_id, dtable_uuid, page_id, tmp_content_file, need_check_static)
        else:
            update_page_design_content_to_path(workspace_id, dtable_uuid, page_id, tmp_page_path, need_check_static)
        # upload
        upload_page_design(repo_id, dtable_uuid, page_id, tmp_page_path, is_dir, username)
    except Exception as e:
        raise e
    finally:
        if is_dir:
            clear_tmp_dir(tmp_page_path)
            clear_tmp_file(tmp_page_path + '.zip')
        else:
            clear_tmp_file(tmp_page_path)
